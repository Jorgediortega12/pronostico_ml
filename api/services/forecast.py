import joblib
from ..config.schemas import ForecastSchema, NewUserModelSchema
import numpy as np
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import func
import pandas as pd
import numpy as np
from ..config.schemas import DemandSearchSchema, CreateOrUpdateUsersModelsSchema, MonthlyInfoSchema, DayBehaviorSchema, CreateTypeYearListSchema, UpdateTypeYearListSchema, UpdateMonthlyTypeSchema, SearchAllMacroeconomicSchema, ForecastTypeMonthSchema, ForecastModelSaveBasedOnYear
from sklearn.preprocessing import StandardScaler
from .analysis import DemandService, MacroeconomicService
from sklearn.impute import SimpleImputer
import joblib
import warnings
from ..other_models.demand import MonthlyDemand, Demand, YearlyDemand, TypeYear
from ..other_models.session import UsersModels, UsersModelsValues
import calendar
from holidays_co import get_colombia_holidays_by_year
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import io
from fastapi import HTTPException
import traceback
import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
import joblib
import warnings
import pandas as pd
import numpy as np
from sklearn.preprocessing import StandardScaler
from sklearn.impute import SimpleImputer
import joblib
import warnings
from ..other_models.demand import MonthlyDemand, Demand, YearlyDemand, TypeYear
warnings.filterwarnings('ignore')

class DemandPredictor:
    def __init__(self, model_path='arima_model/'):
        try:
            """
            Inicializa el predictor cargando los modelos y escaladores necesarios.
            """
            self.model_path = model_path
            
            if model_path.endswith('xgb_model/') or model_path.endswith('rf_model/'):

                self.scaler_x = joblib.load(f'{model_path}annual_scaler_x.pkl')
                self.scaler_y = joblib.load(f'{model_path}annual_scaler_y.pkl')
                self.annual_model = joblib.load(f'{model_path}annual_model.pkl')
                self.monthly_model = joblib.load(f'{model_path}monthly_model.pkl')
                self.imputer = SimpleImputer(strategy='mean')
            
        
            # Obtener las características directamente del scaler
        
                self.annual_features = list(self.scaler_x.feature_names_in_)
            
                # Variables derivadas (para identificarlas al filtrar)
                self.derived_features = {
                    'annual': ['lag_1', 'lag_2', 'lag_3', 'growth_rate',
                                'year_index', 'year_index_norm', 'trend_squared',
                                'growth_momentum', 'ma_2'],
                    'monthly': ['mes', 'mes_sin', 'mes_cos', 'trend',
                                'percentage_lag_12', 'percentage_lag_1', 'dias_laborales',
                                'pct_mean_mes']
                }
                '''
                print("Características requeridas para el modelo anual:")
                print(self.scaler_x.feature_names_in_)
                '''
                metrics=joblib.load(f'{model_path}metrics.pkl')
                self.annual_mape=metrics['annual']['mape']
                self.monthly_mape=metrics['monthly']['mape']
        
            
            elif model_path=='arima_model/':
                self.arima_model = joblib.load(f'{model_path}arima_model.pkl')
                self.mape=joblib.load(f'{model_path}metrics.pkl')
        
            elif model_path.endswith('sarimax_model/'):
                self.sarimax_model = joblib.load(f'{model_path}sarimax_model.pkl')
                self.scaler=joblib.load(f'{model_path}scaler.pkl')
                self.info=joblib.load(f'{model_path}model_info.pkl')
                self.mape=self.info['mape']
        except Exception as e:
            print(f"Error al cargar el modelo: {str(e)}")

    def get_mape(self):
        if self.model_path=='arima_model/':
            return self.mape
        elif self.model_path=='xgb_model/' or self.model_path=='rf_model/':
            mape={'annual':self.annual_mape,'monthly':self.monthly_mape}
            return mape
        elif self.model_path=='sarimax_model/':
            return self.mape
        
    def prepare_annual_features_future(self, historical_data, forecast_years, future_macro_vars):
        """
        Prepara las características para predicciones futuras anuales.
        Versión actualizada con features de tendencia cuadrática.
        """
        required_features = [
            'ipc',
            'Crecimiento de la población (%) anual',
            'Inflación, deflactor del PIB: series vinculadas (% anual)',
            'Comercio de mercaderías (% del PIB)',
            'Importaciones de bienes (balanza de pagos, US$ a precios actuales)',
            'Empleo de tiempo parcial, total (% del total de empleo)',
            'lag_1',
            'lag_2',
            'lag_3',
            'growth_rate',
            'year_index',
            'year_index_norm',
            'trend_squared',
            'growth_momentum',
            'ma_2'
        ]
        
        # 1. Preparar los datos históricos con todas las features
        hist_data = historical_data.copy()
        hist_data['log_total'] = np.log1p(hist_data['total'])
        
        # Calcular year_index para datos históricos
        hist_data['year_index'] = range(len(hist_data))
        hist_data['year_index_norm'] = hist_data['year_index'] / len(hist_data)
        hist_data['trend_squared'] = hist_data['year_index_norm'] ** 2
        hist_data['growth_rate'] = hist_data['log_total'].diff()
        hist_data['growth_momentum'] = hist_data['growth_rate'].rolling(
            window=3, min_periods=1
        ).mean()
        hist_data['ma_2'] = hist_data['log_total'].rolling(
            window=2, min_periods=1
        ).mean()
        
        # 2. Crear DataFrame futuro con las variables macro
        future_data = future_macro_vars.copy()
        
        # 3. Calcular lags usando los datos históricos
        last_log_total = hist_data['log_total'].iloc[-3:]
        future_data['lag_1'] = last_log_total.iloc[-1]
        future_data['lag_2'] = last_log_total.iloc[-2]
        future_data['lag_3'] = last_log_total.iloc[-3]
        
        # 4. Calcular growth rate
        last_growth = hist_data['log_total'].diff().iloc[-1]
        future_data['growth_rate'] = last_growth
        
        # 5. Calcular features de tendencia (NUEVO)
        # El year_index continúa desde donde quedó el histórico
        last_year_index = hist_data['year_index'].iloc[-1]
        total_years = len(hist_data) + forecast_years
        
        future_year_indices = range(last_year_index + 1, last_year_index + 1 + forecast_years)
        future_data['year_index'] = list(future_year_indices)
        future_data['year_index_norm'] = future_data['year_index'] / total_years
        future_data['trend_squared'] = future_data['year_index_norm'] ** 2
        
        # 6. Growth momentum (usar último valor histórico como aproximación)
        last_momentum = hist_data['growth_momentum'].iloc[-1]
        future_data['growth_momentum'] = last_momentum
        
        # 7. MA_2 (usar últimos valores históricos)
        last_ma_2 = hist_data['ma_2'].iloc[-1]
        future_data['ma_2'] = last_ma_2
        
        # 8. Verificar y ordenar columnas
        for feature in required_features:
            if feature not in future_data.columns:
                future_data[feature] = np.nan
        
        # 9. Ordenar columnas según el orden requerido
        X_future = future_data[required_features]
        
        return X_future

    def prepare_single_month_features(self, hist_monthly, current_year, current_month, year_monthly_values, year_total=None):
        """
        Prepara features para un solo mes, similar al método predict_monthly_values del script de entrenamiento.
        """
        # Calcular días laborales para este mes (con fallback si no hay workalendar)
        try:
            from workalendar.america import Colombia
            cal = Colombia()
            first_date = pd.Timestamp(current_year, current_month, 1)
            if current_month == 12:
                last_date = pd.Timestamp(current_year + 1, 1, 1) - pd.Timedelta(days=1)
            else:
                last_date = pd.Timestamp(current_year, current_month + 1, 1) - pd.Timedelta(days=1)
            dias_laborales = cal.get_working_days_delta(first_date.to_pydatetime(), last_date.to_pydatetime())
        except ImportError:
            # Fallback: aproximación simple de días laborales
            days_in_month = pd.Timestamp(current_year, current_month, 1).days_in_month
            # Aproximadamente 22 días laborales por mes (5/7 de los días)
            dias_laborales = int(days_in_month * 5/7)

        # Features básicas
        features = {
            'mes': current_month,
            'mes_sin': np.sin(2 * np.pi * current_month / 12),
            'mes_cos': np.cos(2 * np.pi * current_month / 12),
            'trend': len(hist_monthly) + len(year_monthly_values),  # Índice continuo
            'dias_laborales': dias_laborales,
            'trimestre': ((current_month - 1) // 3 + 1)
        }

        # Lags
        if len(year_monthly_values) > 0:
            # Si ya tenemos predicciones de meses anteriores de este ano
            last_month_percentage = year_monthly_values[-1] / year_total if year_total else 1/12
            features['percentage_lag_1'] = last_month_percentage
        else:
            # Primer mes del ano, usar último mes del ano anterior
            prev_year_data = hist_monthly[hist_monthly['ano'] == current_year - 1]
            if len(prev_year_data) > 0:
                features['percentage_lag_1'] = prev_year_data['percentage'].iloc[-1]
            else:
                features['percentage_lag_1'] = 1/12

        # Lag 12 (mismo mes ano anterior)
        prev_year_same_month = hist_monthly[
            (hist_monthly['ano'] == current_year - 1) &
            (hist_monthly['mes'] == current_month)
        ]
        if len(prev_year_same_month) > 0:
            features['percentage_lag_12'] = prev_year_same_month['percentage'].iloc[0]
        else:
            # Usar promedio del mes
            same_month_data = hist_monthly[hist_monthly['mes'] == current_month]
            features['percentage_lag_12'] = same_month_data['percentage'].mean() if len(same_month_data) > 0 else 1/12

        # Lag 13 (mismo mes hace 2 anos)
        prev_year_2_same_month = hist_monthly[
            (hist_monthly['ano'] == current_year - 2) &
            (hist_monthly['mes'] == current_month)
        ]
        if len(prev_year_2_same_month) > 0:
            features['percentage_lag_13'] = prev_year_2_same_month['percentage'].iloc[0]
        else:
            features['percentage_lag_13'] = features['percentage_lag_12']  # Usar lag_12 como fallback

        # Media móvil del mismo mes (últimos 3 anos)
        same_month_data = hist_monthly[
            (hist_monthly['mes'] == current_month) &
            (hist_monthly['ano'] < current_year) &
            (hist_monthly['ano'] >= max(current_year - 3, hist_monthly['ano'].min()))
        ]['percentage']
        features['ma_same_month_3y'] = same_month_data.mean() if len(same_month_data) > 0 else 1/12

        # Promedio histórico del mes
        hist_data = hist_monthly[
            (hist_monthly['mes'] == current_month) &
            (hist_monthly['ano'] < current_year)
        ]['percentage']
        features['month_historical_avg'] = hist_data.mean() if len(hist_data) > 0 else 1/12

        # Convertir a DataFrame en el orden correcto
        feature_order = ['mes', 'mes_sin', 'mes_cos', 'trend',
                        'percentage_lag_1', 'percentage_lag_12', 'percentage_lag_13',
                        'dias_laborales', 'trimestre',
                        'ma_same_month_3y', 'month_historical_avg']

        return pd.DataFrame([features], columns=feature_order)

    def prepare_year_monthly_features(self, hist_monthly, current_year):
        """
        Prepara features para TODOS los meses de un ano de una vez.
        Versión actualizada 2025 con nuevas features (pct_mean_mes).
        """
        months = range(1, 13)

        # Calcular días laborales para cada mes
        try:
            from workalendar.america import Colombia
            cal = Colombia()
            dias_lab = []
            for month in months:
                first_date = pd.Timestamp(current_year, month, 1)
                if month == 12:
                    last_date = pd.Timestamp(current_year + 1, 1, 1) - pd.Timedelta(days=1)
                else:
                    last_date = pd.Timestamp(current_year, month + 1, 1) - pd.Timedelta(days=1)
                dias_laborales = cal.get_working_days_delta(first_date.to_pydatetime(), last_date.to_pydatetime())
                dias_lab.append(dias_laborales)
        except ImportError:
            # Fallback: aproximación simple
            dias_lab = []
            for month in months:
                days_in_month = pd.Timestamp(current_year, month, 1).days_in_month
                dias_lab.append(int(days_in_month * 5/7))

        # Calcular year_idx como en el script de entrenamiento
        last_year_in_data = hist_monthly['ano'].max()
        last_idx = hist_monthly[hist_monthly['ano'] == last_year_in_data].index[-1]
        years_ahead = current_year - last_year_in_data
        year_idx = last_idx + 1 + (years_ahead - 1) * 12

        # Features básicas
        X_pred = pd.DataFrame({
            'mes': months,
            'mes_sin': np.sin(2 * np.pi * np.array(months)/12),
            'mes_cos': np.cos(2 * np.pi * np.array(months)/12),
            'trend': range(year_idx, year_idx + 12),
            'dias_laborales': dias_lab
        })

        # Lags
        if current_year > hist_monthly['ano'].min():
            # Datos del ano anterior
            prev_year_data = hist_monthly[hist_monthly['ano'] == current_year - 1]

            if len(prev_year_data) > 0:
                # lag_12: usar ano anterior completo
                X_pred['percentage_lag_12'] = prev_year_data['percentage'].values
            else:
                # Fallback: promedio histórico por mes
                hist_avg_by_month = []
                for m in months:
                    hist_data = hist_monthly[hist_monthly['mes'] == m]['percentage']
                    hist_avg_by_month.append(hist_data.mean() if len(hist_data) > 0 else 1/12)
                X_pred['percentage_lag_12'] = hist_avg_by_month

            # lag_1: usar último mes del ano anterior para TODOS los meses
            if len(prev_year_data) > 0:
                last_month_prev_year = prev_year_data['percentage'].iloc[-1]
                X_pred['percentage_lag_1'] = last_month_prev_year
            else:
                X_pred['percentage_lag_1'] = 1/12

            # pct_mean_mes: Promedio móvil de 3 anos del mismo mes
            # Replica: df.groupby('mes')['percentage'].transform(lambda s: s.shift(1).rolling(3, min_periods=1).mean())
            pct_mean_values = []
            for m in months:
                # Obtener porcentajes históricos del mismo mes
                same_month_data = hist_monthly[
                    (hist_monthly['mes'] == m) &
                    (hist_monthly['ano'] < current_year)
                ].sort_values('ano')['percentage']

                if len(same_month_data) > 0:
                    # Tomar los últimos 3 valores disponibles y promediar
                    recent_values = same_month_data.tail(3)
                    pct_mean_values.append(recent_values.mean())
                else:
                    pct_mean_values.append(1/12)

            X_pred['pct_mean_mes'] = pct_mean_values

        else:
            # Primer ano: usar valores por defecto
            X_pred['percentage_lag_1'] = 1/12
            X_pred['percentage_lag_12'] = 1/12
            X_pred['pct_mean_mes'] = 1/12

        # Orden correcto: mes, mes_sin, mes_cos, trend, percentage_lag_12, percentage_lag_1, dias_laborales, pct_mean_mes
        return X_pred[['mes', 'mes_sin', 'mes_cos', 'trend', 'percentage_lag_12', 'percentage_lag_1', 'dias_laborales', 'pct_mean_mes']]

    def prepare_monthly_features_future(self, historical_monthly_data, num_months):
        """
        Prepara las características para predicciones futuras mensuales.
        Versión actualizada 2025 con nuevas features (pct_mean_mes).
        """
        # Calcular el porcentaje mensual en los datos históricos
        hist_data = historical_monthly_data.copy()
        hist_data['yearly_total'] = hist_data.groupby('ano')['total'].transform('sum')
        hist_data['percentage'] = hist_data['total'] / hist_data['yearly_total']

        # Crear DataFrame para meses futuros
        last_date = pd.to_datetime(hist_data['fecha'].max())
        future_dates = pd.date_range(
            start=last_date + pd.DateOffset(months=1),
            periods=num_months,
            freq='M'
        )

        future_df = pd.DataFrame({
            'fecha': future_dates,
            'ano': future_dates.year,
            'mes': future_dates.month,
            'dias_laborales': 22  # Valor por defecto
        })

        # Features básicas
        future_df['mes_sin'] = np.sin(2 * np.pi * future_df['mes']/12)
        future_df['mes_cos'] = np.cos(2 * np.pi * future_df['mes']/12)
        future_df['trend'] = range(
            len(hist_data),
            len(hist_data) + len(future_df)
        )

        # Combinar datos históricos y futuros para calcular lags
        df_combined = pd.concat(
            [hist_data[['fecha', 'mes', 'ano', 'percentage']], future_df],
            ignore_index=True
        )
        df_combined = df_combined.sort_values('fecha')

        # Calcular lags
        df_combined['percentage_lag_1'] = df_combined['percentage'].shift(1)
        df_combined['percentage_lag_12'] = df_combined['percentage'].shift(12)

        # Calcular pct_mean_mes: Promedio móvil de 3 anos del mismo mes
        pct_mean_values = []
        for idx, row in df_combined.tail(num_months).iterrows():
            mes_actual = row['mes']
            año_actual = row['ano']

            # Obtener porcentajes históricos del mismo mes
            same_month_data = hist_data[
                (hist_data['mes'] == mes_actual) &
                (hist_data['ano'] < año_actual)
            ].sort_values('ano')['percentage']

            if len(same_month_data) > 0:
                # Tomar los últimos 3 valores disponibles y promediar
                recent_values = same_month_data.tail(3)
                pct_mean_values.append(recent_values.mean())
            else:
                # Fallback: distribución uniforme
                pct_mean_values.append(1/12)

        # Obtener solo las filas futuras y agregar las features calculadas
        future_df['percentage_lag_1'] = df_combined['percentage_lag_1'].tail(num_months).values
        future_df['percentage_lag_12'] = df_combined['percentage_lag_12'].tail(num_months).values
        future_df['pct_mean_mes'] = pct_mean_values

        # Rellenar NaN con valores por defecto (para casos donde no hay suficiente historia)
        future_df['percentage_lag_1'].fillna(1/12, inplace=True)
        future_df['percentage_lag_12'].fillna(1/12, inplace=True)
        future_df['pct_mean_mes'].fillna(1/12, inplace=True)

        # Orden correcto de columnas: mes, mes_sin, mes_cos, trend, percentage_lag_12, percentage_lag_1, dias_laborales, pct_mean_mes
        features = ['mes', 'mes_sin', 'mes_cos', 'trend',
                    'percentage_lag_12', 'percentage_lag_1', 'dias_laborales',
                    'pct_mean_mes']

        return future_df[features]
    
    def filter_features(self, df, model_name):
        """
        Filtra el DataFrame para mantener solo las variables necesarias para el modelo especificado.
        
        Args:
            df (pd.DataFrame): DataFrame con todas las variables (derivadas y macro)
            model_name (str): Nombre del modelo ('annual_model' o 'monthly_model')
        
        Returns:
            pd.DataFrame: DataFrame filtrado con solo las variables necesarias
        """
        if model_name == 'annual_model':
            required_features = self.annual_features
        elif model_name == 'monthly_model':
            required_features = self.derived_features['monthly']
        else:
            raise ValueError(f"Modelo '{model_name}' no soportado")

        # Verificar que todas las variables requeridas estén presentes
        missing_features = set(required_features) - set(df.columns)
        if missing_features:
            raise ValueError(f"Faltan las siguientes variables requeridas: {missing_features}")
        
        # Mantener solo las columnas necesarias en el orden correcto
        filtered_df = df[required_features].copy()
        
        return filtered_df
    
    def remove_atypical_years(self, df, atypical_years, year_column='ano'):
        """
        Removes rows from a DataFrame where the year is in the specified atypical_years array.
        
        Parameters:
        -----------
        df : pandas.DataFrame
            The input DataFrame containing the year column
        atypical_years : list or array-like
            List of years to remove from the DataFrame
        year_column : str, default='ano'
            Name of the column containing the years
            
        Returns:
        --------
        pandas.DataFrame
            A new DataFrame with the atypical years removed
            
        Raises:
        -------
        ValueError
            If the year_column doesn't exist in the DataFrame
            If atypical_years is empty or contains non-numeric values
        """
        try:
            # Input validation
            if df.index.name == year_column:
                mask = ~df.index.isin(atypical_years)
                filtered_df = df[mask].copy()

            elif year_column not in df.columns:
                raise ValueError(f"Column '{year_column}' not found in DataFrame. "
                                f"Available columns are: {', '.join(df.columns)}")
            else:
                # Create mask for filtering
                mask = ~df[year_column].isin(atypical_years)
                
                # Apply filter and create new DataFrame
                filtered_df = df[mask].copy()
            
            # Log information about removed data
            removed_count = len(df) - len(filtered_df)
            print(f"Removed {removed_count} rows containing years: {sorted(atypical_years)}")
            print(f"Remaining data shape: {filtered_df.shape}")
        except Exception as e:
            print("Error durante el filtrado de anos atipicos", str(e))
            raise e
        return filtered_df

    def predict_future_arima(self,df, forecast_periods=12):
        """
        Realiza predicciones usando un modelo ARIMA previamente entrenado.
        
        Args:
            df (pd.DataFrame): DataFrame con columnas ['ano', 'mes', 'total']
            forecast_periods (int): Número de meses a predecir
            
        Returns:
            pd.DataFrame: Predicciones con fechas futuras
        """
       
        try:
            # 1. Cargar el modelo
            model = self.arima_model
            
            
            # 3. Hacer predicciones
            forecast = model.forecast(steps=forecast_periods)
            
            # 4. Crear fechas futuras
            last_date = df['fecha'].max()
            future_dates = pd.date_range(
                start=last_date + pd.DateOffset(months=1),
                periods=forecast_periods,
                freq='M'
            )
            
            # 5. Crear DataFrame de resultados
            predictions = pd.DataFrame({
                'fecha': future_dates,
                'ano': future_dates.year,
                'mes': future_dates.month,
                'prediccion': forecast
            })
            
            return predictions
            
        except Exception as e:
            print(f"Error al realizar predicciones: {str(e)}")
            raise e
        

    def predict_future_xgrf(self, historical_annual_data, historical_monthly_data, 
                    forecast_years=1, future_macro_vars=None, atypical_years=None):
        """
        Realiza predicciones futuras para los próximos anos de forma secuencial.
        """
        if future_macro_vars is None:
            raise ValueError("Debes proporcionar las variables macroeconómicas futuras")
            
        try:
            # Preparar datos anuales futuros
            # historical_annual_data = self.remove_atypical_years(historical_annual_data, atypical_years)
            # print("con annual no es")
            print(historical_annual_data.head())

            
            # Preparar datos históricos con log
            hist_data = historical_annual_data.copy()
            hist_data['log_total'] = np.log1p(hist_data['total'])
            
            # Variables requeridas
            required_features = [
                'ipc',
                'Crecimiento de la población (%) anual',
                'Inflación, deflactor del PIB: series vinculadas (% anual)',
                'Comercio de mercaderías (% del PIB)',
                'Importaciones de bienes (balanza de pagos, US$ a precios actuales)',
                'Empleo de tiempo parcial, total (% del total de empleo)',
                'lag_1',
                'lag_2',
                'lag_3',
                'growth_rate',
                'year_index',
                'year_index_norm',
                'trend_squared',
                'growth_momentum',
                'ma_2'
            ]
            
            # Inicializar listas para predicciones
            annual_predictions = []
            
            # Obtener los últimos valores para inicializar los lags
            last_log_totals = hist_data['log_total'].values[-3:].tolist()  # Últimos 3 valores como lista
            last_growth = hist_data['log_total'].diff().iloc[-1]
            
            # Realizar predicciones secuencialmente
            for year_idx in range(forecast_years):
                print(f"\n=== Prediciendo ano {future_macro_vars.index[year_idx]} ===")
                
                # Crear DataFrame para el ano actual
                current_year_data = pd.DataFrame(
                    future_macro_vars.iloc[year_idx]).T
                
                # ===== CALCULAR FEATURES DE LAGS =====
                current_year_data['lag_1'] = last_log_totals[-1]
                current_year_data['lag_2'] = last_log_totals[-2]
                current_year_data['lag_3'] = last_log_totals[-3]
                current_year_data['growth_rate'] = last_growth
                
                # ===== CALCULAR FEATURES DE TENDENCIA (NUEVO) =====
                # Calcular year_index: continúa desde donde quedó el histórico
                current_year_number = future_macro_vars.index[year_idx]
                years_from_start = current_year_number - hist_data.index[0]  # Años desde el inicio del histórico
                current_year_data['year_index'] = years_from_start
                
                # Calcular year_index_norm: normalizado respecto al total de anos disponibles
                # (histórico + anos futuros ya predichos + ano actual)
                total_years_so_far = len(hist_data) + year_idx + 1
                current_year_data['year_index_norm'] = current_year_data['year_index'] / total_years_so_far
                
                # Calcular trend_squared
                current_year_data['trend_squared'] = current_year_data['year_index_norm'] ** 2
                
                # Calcular growth_momentum: usar último valor disponible
                if year_idx == 0:
                    # Primer ano futuro: usar último momentum del histórico
                    last_growth_values = hist_data['log_total'].diff().tail(3)
                    current_year_data['growth_momentum'] = last_growth_values.mean()
                else:
                    # Años subsiguientes: promedio de últimos crecimientos (incluyendo predichos)
                    current_year_data['growth_momentum'] = last_growth
                
                # Calcular ma_2: promedio móvil de últimos 2 log_totals
                if year_idx == 0:
                    # Primer ano: usar último valor histórico
                    last_two = hist_data['log_total'].tail(2).mean()
                    current_year_data['ma_2'] = last_two
                else:
                    # Años subsiguientes: promedio de últimos dos valores (incluyendo predichos)
                    current_year_data['ma_2'] = (last_log_totals[-1] + last_log_totals[-2]) / 2
                
                # ===== LOG DE DEBUG =====
                print(f"Lags para ano {future_macro_vars.index[year_idx]}:")
                print(f"  lag_1: {current_year_data['lag_1'].values[0]:.4f}")
                print(f"  lag_2: {current_year_data['lag_2'].values[0]:.4f}")
                print(f"  lag_3: {current_year_data['lag_3'].values[0]:.4f}")
                print(f"  growth_rate: {current_year_data['growth_rate'].values[0]:.4f}")
                print(f"Features de tendencia:")
                print(f"  year_index: {current_year_data['year_index'].values[0]:.0f}")
                print(f"  year_index_norm: {current_year_data['year_index_norm'].values[0]:.4f}")
                print(f"  trend_squared: {current_year_data['trend_squared'].values[0]:.4f}")
                print(f"  growth_momentum: {current_year_data['growth_momentum'].values[0]:.4f}")
                print(f"  ma_2: {current_year_data['ma_2'].values[0]:.4f}")
                
                # Verificar y ordenar columnas
                for feature in required_features:
                    if feature not in current_year_data.columns:
                        current_year_data[feature] = np.nan
                
                # Ordenar columnas según el orden requerido
                X_current = current_year_data[required_features]
                
                # Filtrar características
                X_current_filtered = self.filter_features(X_current, 'annual_model')
                
                # Realizar predicción
                X_current_scaled = self.scaler_x.transform(X_current_filtered)
                annual_pred_scaled = self.annual_model.predict(X_current_scaled)
                annual_pred_value = np.expm1(
                    self.scaler_y.inverse_transform(annual_pred_scaled.reshape(-1, 1))
                )[0][0]
                
                print(f"Predicción para ano {future_macro_vars.index[year_idx]}: {annual_pred_value:,.0f}")
                prev_value = float(np.expm1(last_log_totals[-1]))  # último total (histórico o predicho)
                if prev_value > 0:
                    implied_growth = (annual_pred_value - prev_value) / prev_value
                    if implied_growth < -0.30:  # caída interanual > 30% = sospechosa
                        hist_log_growth = hist_data['log_total'].diff().dropna()
                        avg_log_growth = float(hist_log_growth.tail(5).mean()) if len(hist_log_growth) else 0.0
                        avg_log_growth = max(0.0, avg_log_growth)  # no arrastrar tendencias negativas espurias
                        sane_value = float(np.expm1(last_log_totals[-1] + avg_log_growth))
                        print(f"⚠️ Predicción degenerada para {future_macro_vars.index[year_idx]}: "
                              f"{annual_pred_value:,.0f} (growth {implied_growth:.1%}). "
                              f"Corrigiendo a {sane_value:,.0f} (growth log medio {avg_log_growth:.4f}).")
                        annual_pred_value = sane_value

                # Guardar predicción
                annual_predictions.append(annual_pred_value)
                
                # IMPORTANTE: Actualizar lags para la siguiente predicción
                # Calcular el log de la predicción actual
                current_log_pred = np.log1p(annual_pred_value)
                
                # Actualizar la lista de logs (eliminar el más antiguo y agregar el nuevo)
                last_log_totals.pop(0)  # Eliminar el más antiguo
                last_log_totals.append(current_log_pred)  # Agregar el nuevo
                
                # Actualizar growth rate
                last_growth = current_log_pred - last_log_totals[-2]  # Diferencia con el valor anterior
                
                print(f"Actualizando para siguiente predicción:")
                print(f"  Nuevos últimos 3 logs: {[f'{x:.4f}' for x in last_log_totals]}")
                print(f"  Nuevo growth_rate: {last_growth:.4f}")
            
            # Convertir a array numpy
            annual_pred = np.array(annual_predictions).reshape(-1, 1)
            
            # Preparar datos mensuales futuros
            X_future_monthly = self.prepare_monthly_features_future(
                historical_monthly_data,
                forecast_years * 12
            )
            
            X_future_monthly = self.filter_features(X_future_monthly, 'monthly_model')
            
            # Predecir porcentajes mensuales EN LOTE (como en el script de entrenamiento)
            monthly_predictions = []

            # Obtener datos históricos de porcentajes
            hist_monthly = historical_monthly_data.copy()
            hist_monthly['yearly_total'] = hist_monthly.groupby('ano')['total'].transform('sum')
            hist_monthly['percentage'] = hist_monthly['total'] / hist_monthly['yearly_total']

            for year_idx in range(forecast_years):
                year_total = annual_pred[year_idx][0]
                current_year = hist_monthly['ano'].max() + 1 + year_idx

                # Preparar features para TODO EL AÑO de una vez (como en training script)
                year_features = self.prepare_year_monthly_features(
                    hist_monthly, current_year
                )

                # Filtrar features según el modelo
                year_features_filtered = self.filter_features(year_features, 'monthly_model')

                # Predecir porcentajes para TODO EL AÑO
                year_percentages = self.monthly_model.predict(year_features_filtered)

                # Normalizar porcentajes para que sumen 1
                year_percentages = np.maximum(year_percentages, 0)  # Evitar negativos
                year_percentages = year_percentages / year_percentages.sum()

                # Calcular valores mensuales
                year_monthly_values = year_total * year_percentages
                monthly_predictions.extend(year_monthly_values)

                # Agregar este ano predicho a los datos históricos para el siguiente ano
                for month in range(1, 13):
                    new_row = pd.DataFrame({
                        'ano': [current_year],
                        'mes': [month],
                        'total': [year_monthly_values[month-1]],
                        'fecha': [pd.Timestamp(current_year, month, 1)],
                        'yearly_total': [year_total],
                        'percentage': [year_percentages[month-1]]
                    })
                    hist_monthly = pd.concat([hist_monthly, new_row], ignore_index=True)
            
            # Crear DataFrame de resultados
            last_date = pd.to_datetime(historical_monthly_data['fecha'].max())
            future_dates = pd.date_range(
                start=last_date + pd.DateOffset(months=1),
                periods=len(monthly_predictions),
                freq='M'
            )
            
            results = pd.DataFrame({
                'fecha': future_dates,
                'prediccion': monthly_predictions
            })
            
            return annual_pred, results
            
        except Exception as e:
            print("Error durante la prediccion de xgrf", str(e))
            raise e

    def predict_future_sarimax(self,df_demanda,future_macro_vars, forecast_periods=12,atypical_years=None):
        """
        Realiza predicciones usando el modelo SARIMAX guardado
        
        Args:
            df (pd.DataFrame): DataFrame con columnas ano, mes, total y variables macro
            forecast_periods (int): Períodos a predecir
        """
        df_demanda=self.remove_atypical_years(df_demanda, atypical_years)
        # Cargar modelo y componentes
        model = self.sarimax_model
        scaler = self.scaler
        info = self.info
        vars_macro = info['vars_macro']
         # Asegurar que df_demanda tiene fecha como índice
        if 'fecha' in df_demanda.columns:
            df_demanda = df_demanda.set_index('fecha')
    
        historical_data = pd.merge(
            df_demanda,
            future_macro_vars,
            left_on='ano',
            right_index=True,
            how='left'
        )
        
        y_log = np.log1p(historical_data['total'])
        exog_vars = historical_data[vars_macro]
        exog_scaled = scaler.transform(exog_vars)
        
        # Crear fechas futuras correctamente
        last_date = df_demanda.index[-1]
        future_dates = pd.date_range(
            start=last_date + pd.DateOffset(months=1),
            periods=forecast_periods,
            freq='M'
        )
        
        future_exog = pd.DataFrame(index=future_dates)
        
        for column in vars_macro:
            future_exog[column] = future_macro_vars[column].iloc[-1]
        
        future_exog_scaled = scaler.transform(future_exog)
        
        forecast_log = model.forecast(steps=forecast_periods, exog=future_exog_scaled)
        forecast = np.expm1(forecast_log)

        print("La predicción de SARIMAX fue exitosa")
        
        return pd.DataFrame({
            'fecha': future_dates,
            'prediccion': forecast
        })

def check_current_year(start_date, end_date):
    """
    Verifica si la fecha actual está dentro del rango de fechas.
    Si esta dentro del rango levanta un error
    """
    current_date = pd.to_datetime('today').date()
    start_date = pd.to_datetime(start_date).date()
    end_date = pd.to_datetime(end_date).date()
    if start_date <= current_date <= end_date:
        raise ValueError("La fecha actual está dentro del rango de fechas. No se puede seguir con la predicción.")


#Start of endpoints 
class ForecastService:
    def predict(self, search: ForecastSchema, db: Session):
        try:
            # Initialize updated_variables early to avoid reference errors in exception handlers
            updated_variables = {}
            
            demand = DemandService()
            start_date = pd.to_datetime(search.start_date)
            end_date = pd.to_datetime(search.end_date)

            # Query monthly demand data up to the start date year
            # Query monthly demand data up to the start date (using both year and month)
            historic_demands = db.query(MonthlyDemand.value, MonthlyDemand.year, MonthlyDemand.month)\
                .filter(
                    # Exclude data from after start date (including specific month)
                    ((MonthlyDemand.year < start_date.year) |
                     ((MonthlyDemand.year == start_date.year) & (MonthlyDemand.month < start_date.month)))
                )\
                .order_by(MonthlyDemand.year, MonthlyDemand.month)\
                .all()
                
            # Create separate arrays for demands and dates while preserving their order
            demand_values_for_table = []
            date_values_for_table = []

            for record in historic_demands:
                demand_values_for_table.append(record.value)
                # Format date as 'YYYY-MM-01'
                date_values_for_table.append(f"{record.year}-{record.month:02d}-01")

  
            historic_demands_for_table = historic_demands


            # CAMBIO 1: Eliminamos la validación de fecha inicial
            # Ya no verificamos si start_date > last_date_ym
            
            # CAMBIO 2: Calcular la fecha límite para datos históricos (último día del mes anterior)
            # Si start_date es 2025-01-01, queremos 2024-12-31 (último día de diciembre)
            historical_limit_date = start_date.replace(day=1) - pd.DateOffset(days=1)
            historical_limit_date_str = historical_limit_date.strftime('%Y-%m-%d')
            
            # Obtener el rango completo de datos disponibles para determinar el inicio
            year_range = demand.get_first_and_last_date(db)
            print("year_range", year_range)
            first_date = year_range["first_date"].strftime('%Y-%m-%d')
            # Usamos la fecha límite calculada en lugar de la última fecha disponible
            last_date = historical_limit_date_str
            first_year = int(pd.to_datetime(first_date).year)
            last_year = int(pd.to_datetime(last_date).year)
            
            print(f"Usando datos históricos desde {first_date} hasta {last_date}")
            print(f"Predicción solicitada desde {search.start_date} hasta {search.end_date}")
            
            try:
                print("Creando MacroeconomicService...")
                macroeconomic = MacroeconomicService()
                
                # Calculate forecast years based on difference between years
                print(f"Calculando anos de pronóstico. start_date: {start_date}, end_date: {end_date}")
                forecast_years = end_date.year - start_date.year + 1
                years_list = []
                for i in range(forecast_years):
                    years_list.append(start_date.year + i)
                start_year = min(years_list)
                print(f"Años a pronosticar: {years_list}, start_year: {start_year}")

                # Crear esquema de búsqueda
                print(f"Creando esquema de búsqueda con ano_inicio={start_year}, ano_fin={end_date.year}")
                try:
                    macro_Search = SearchAllMacroeconomicSchema(
                        ano_inicio=str(start_year),
                        ano_fin=str(end_date.year)
                    )
                    print(f"Esquema creado: {macro_Search}")
                except Exception as e:
                    print(f"Error al crear el esquema SearchAllMacroeconomicSchema: {str(e)}")
                    return {"message": f"Error al crear esquema de búsqueda: {str(e)}", "status": "error"}
                
                # Obtener variables macroeconómicas
                print("Obteniendo variables macroeconómicas...")
                macro_variables = macroeconomic.get_all_variables(db, macro_Search)
                
                # Verificar si hay variables disponibles
                if not macro_variables:
                    return {
                        "message": f"No hay variables macroeconómicas disponibles para el período {start_year} - {end_date.year}. Por favor, utilice un rango de fechas para el que existan datos históricos completos.",
                        "status": "error"
                    }
                print("macro_variables", macro_variables)
                
            except Exception as e:
                print(f"Error inesperado en el bloque de variables macroeconómicas: {str(e)}")
                return {"message": f"Error al procesar variables macroeconómicas: {str(e)}", "status": "error"}
                
            # Define the mapping between API variable names and model feature names
            macroeconomic_variables = {
                'ipc': 'ipc',
                'Crecimiento': 'Crecimiento de la población (%) anual',
                'Inflación': 'Inflación, deflactor del PIB: series vinculadas (% anual)',
                'Comercio': 'Comercio de mercaderías (% del PIB)',
                'Importaciones': 'Importaciones de bienes (balanza de pagos, US$ a precios actuales)',
                'Empleo': 'Empleo de tiempo parcial, total (% del total de empleo)'
            }
            
            # Verificar que todas las variables necesarias existan para TODOS los anos
            missing_vars = []
            for api_name, model_name in macroeconomic_variables.items():
                print(f"Verificando variable: {api_name} -> {model_name}")
                
                # Verificar si la variable existe
                if api_name not in macro_variables:
                    print(f"Error: La variable macroeconómica '{api_name}' no fue encontrada.")
                    missing_vars.append(api_name)
                    continue
                
                # Verificar que haya suficientes valores para cada ano
                variable_values = macro_variables[api_name]
                if len(variable_values) < len(years_list):
                    print(f"Error: La variable '{api_name}' no tiene datos para todos los anos requeridos.")
                    print(f"  - Años requeridos: {len(years_list)}, Datos disponibles: {len(variable_values)}")
                    missing_vars.append(f"{api_name} (faltan datos para algunos anos)")
                    continue
                
                # Guardar los valores sin rellenar
                updated_variables[model_name] = variable_values[:len(years_list)]
            
            # Si falta alguna variable, no continuar con la predicción
            if missing_vars:
                return {
                    "message": f"No se puede realizar la predicción porque faltan las siguientes variables macroeconómicas o no están completas para todo el período: {', '.join(missing_vars)}. Por favor, complete estos datos antes de continuar.",
                    "status": "error"
                }
                
            print("updated_variables", updated_variables)

            # Verificar que updated_variables tenga todas las variables necesarias
            expected_vars = set(macroeconomic_variables.values())
            actual_vars = set(updated_variables.keys())
            if expected_vars != actual_vars:
                missing = expected_vars - actual_vars
                return {
                    "message": f"Faltan las siguientes variables macroeconómicas requeridas: {', '.join(missing)}",
                    "status": "error"
                }

            print("Creando esquemas de búsqueda para datos históricos...")
            try:
                search_schema_annual = DemandSearchSchema(
                    tipo=3,
                    ano_inicio=str(first_year),
                    ano_fin=str(last_year)
                )
                search_schema_monthly = DemandSearchSchema(
                    tipo=2,
                    fecha_inicio=str(first_date),
                    fecha_fin=str(last_date)
                )
                print("Esquemas creados correctamente")
            except Exception as e:
                print(f"Error al crear los esquemas de búsqueda: {str(e)}")
                return {"message": f"Error al preparar los esquemas de búsqueda: {str(e)}", "status": "error"}
                
            try:
                print("Obteniendo anos atípicos...")
                atypical_years = db.query(TypeYear.year).filter_by(type="ATIPICO").all()
                atypical_years = [year[0] for year in atypical_years]
                print("atypical_years", atypical_years)
            except Exception as e:
                print(f"Error al obtener anos atípicos: {str(e)}")
                return {"message": f"Error al obtener anos atípicos: {str(e)}", "status": "error"}
            
            # Obtener datos anuales y mensuales
            try:
                print("Obteniendo datos anuales históricos...")
                annual_data_raw = demand.get_demands(db, search_schema_annual)
                print(f"Datos anuales recibidos: {type(annual_data_raw)}")
                
                print("Obteniendo datos mensuales históricos...")
                monthly_data_raw = demand.get_demands(db, search_schema_monthly)
                print(f"Datos mensuales recibidos: {type(monthly_data_raw)}")
            except Exception as e:
                print(f"Error al obtener datos históricos: {str(e)}")
                return {"message": f"Error al obtener datos históricos: {str(e)}", "status": "error"}
            
            # Verificar que los datos no estén vacíos
            if not annual_data_raw or not annual_data_raw.get("lista_demandas") or not annual_data_raw.get("lista_fechas"):
                raise ValueError("No se encontraron datos anuales para el periodo solicitado")
            
            if not monthly_data_raw or not monthly_data_raw.get("lista_demandas") or not monthly_data_raw.get("lista_fechas"):
                raise ValueError("No se encontraron datos mensuales para el periodo solicitado")
            
            # Formatear datos anuales
            annual_data = pd.DataFrame({
                'total': annual_data_raw["lista_demandas"],
                'ano': annual_data_raw["lista_fechas"]
            }, index=annual_data_raw["lista_fechas"])
            
            # Formatear datos mensuales
            dates = pd.to_datetime(monthly_data_raw["lista_fechas"])
            monthly_data = pd.DataFrame({
                'ano': dates.year,
                'mes': dates.month,
                'total': monthly_data_raw["lista_demandas"],
                'fecha': dates
            })
            
            print("\nRevisión de datos cargados:")
            print("Datos anuales:", annual_data.head())
            print("Datos mensuales:", monthly_data.head())
            
            # Asegurarse que los datos mensuales tengan el formato correcto
            monthly_data['total'] = monthly_data['total'].astype(float)

            print("years_list", years_list)
            

            # Preparar totales anuales historicos ordenados (anio mas viejo primero)
            historical_annual_totals = []
            try:
                historical_years = [int(y) for y in annual_data_raw["lista_fechas"]]
                historical_totals = annual_data_raw["lista_demandas"]
                historical_pairs = sorted(zip(historical_years, historical_totals), key=lambda x: x[0])
                historical_annual_totals = [float(total) for _, total in historical_pairs]
            except Exception as e:
                print(f"Error al preparar totales anuales historicos: {str(e)}")
                historical_annual_totals = [float(total) for total in annual_data_raw.get("lista_demandas", [])]

            # Crear DataFrame para variables macroeconómicas futuras
            future_macro_vars = pd.DataFrame(
                updated_variables,
                index=years_list
            )
            
            yearly_demands = db.query(YearlyDemand).all()
            years = {}
            for yearly_demand in yearly_demands:  # Cambié 'demand' por 'yearly_demand'
                years[yearly_demand.year] = yearly_demand.demand

            # IMPORTANTE: Guardar los anos que realmente queremos predecir
            years_to_predict = years_list.copy()
            
            # Guardar el número de anos que queremos predecir ANTES de llamar al modelo
            num_years_to_predict = len(years_to_predict)
                
            if search.id == 1 or search.id == 2:
                if search.id == 1:
                    model_path_name = 'api/models/rf_model/'
                else:   
                    model_path_name = 'api/models/xgb_model/'

                predictor = DemandPredictor(model_path=model_path_name)
            
                # Realizar predicción
                annual_predictions, monthly_predictions = predictor.predict_future_xgrf(
                    historical_annual_data=annual_data,
                    historical_monthly_data=monthly_data,
                    forecast_years=num_years_to_predict,  # Usar el número guardado
                    future_macro_vars=future_macro_vars, 
                    atypical_years=atypical_years
                )
                
                # VALIDACIÓN: Verificar que tengamos el número correcto de predicciones
                if len(annual_predictions) != num_years_to_predict:
                    print(f"ADVERTENCIA: Se esperaban {num_years_to_predict} predicciones pero se obtuvieron {len(annual_predictions)}")
                    # Si hay menos predicciones, podríamos necesitar ajustar o lanzar un error
                    if len(annual_predictions) < num_years_to_predict:
                        raise ValueError(f"El modelo devolvió menos predicciones ({len(annual_predictions)}) que los anos solicitados ({num_years_to_predict})")

                # Calcular crecimiento usando los anos correctos
                for i, year in enumerate(years_to_predict):
                    if i < len(annual_predictions):
                        years[year] = int(annual_predictions[i])
                    else:
                        print(f"ADVERTENCIA: No hay predicción para el ano {year}")
                        
                print("years", years)

                years_list_full = sorted(list(years.keys()))
                print("years_list_full", years_list_full)
                growth = {}
                for i in range(1, len(years_list_full)):
                    if years_list_full[i-1] in years and years_list_full[i] in years:
                        growth[years_list_full[i]] = (years[years_list_full[i]] - years[years_list_full[i-1]]) / years[years_list_full[i-1]]
                print("growth", growth)
                
                # Convertir predicciones a lista simple
                if isinstance(annual_predictions, np.ndarray):
                    annual_predictions = annual_predictions.tolist()
                    if len(annual_predictions) > 0 and isinstance(annual_predictions[0], list):
                        annual_predictions = [pred[0] for pred in annual_predictions]
                print("annual_predictions", annual_predictions)

            elif search.id == 3:
                try:
                    model_path_name = 'api/models/sarimax_model/'
                    predictor = DemandPredictor(model_path=model_path_name)
                    
                    # Para SARIMAX, calcular el número de períodos mensuales
                    forecast_periods = num_years_to_predict * 12
                    
                    monthly_predictions = predictor.predict_future_sarimax(
                        monthly_data,
                        future_macro_vars,
                        atypical_years=atypical_years, 
                        forecast_periods=forecast_periods
                    )
                    print("monthly_predictions", monthly_predictions)
                    
                    # Verificar que monthly_predictions contiene los datos esperados
                    if 'prediccion' not in monthly_predictions or 'fecha' not in monthly_predictions:
                        raise ValueError("La predicción SARIMAX no retornó los campos esperados")
                        
                    # Agrupar predicciones mensuales por ano
                    annual_predictions = monthly_predictions['prediccion'].groupby(monthly_predictions['fecha'].dt.year).sum().tolist()
                    
                    # VALIDACIÓN: Verificar que tengamos el número correcto de predicciones anuales
                    if len(annual_predictions) != num_years_to_predict:
                        print(f"ADVERTENCIA SARIMAX: Se esperaban {num_years_to_predict} predicciones anuales pero se obtuvieron {len(annual_predictions)}")
                    
                    # Calcular crecimiento
                    for i, year in enumerate(years_to_predict):
                        if i < len(annual_predictions):
                            years[year] = int(annual_predictions[i])
                            
                    years_list_full = sorted(list(years.keys()))
                    growth = {}
                    for i in range(1, len(years_list_full)):
                        if years_list_full[i-1] in years and years_list_full[i] in years:
                            growth[years_list_full[i]] = (years[years_list_full[i]] - years[years_list_full[i-1]]) / years[years_list_full[i-1]]
                    print("growth", growth)

                except Exception as e:
                    print("Error en la predicción SARIMAX", str(e))
                    raise ValueError(f"Error en la predicción SARIMAX: {str(e)}")
            else:
                raise ValueError(f"Modelo de predicción no soportado: id={search.id}")

            # CAMBIO 3 y 4: Filtrar resultados para el período exacto solicitado y obtener valores reales
            # Crear lista de fechas mensuales solicitadas
            requested_dates = pd.date_range(
                start=start_date,
                end=end_date,
                freq='M'
            )
            
            # Filtrar predicciones mensuales para el período solicitado
            monthly_predictions_df = pd.DataFrame({
                'fecha': monthly_predictions['fecha'],
                'prediccion': monthly_predictions['prediccion']
            })
            
            # Filtrar solo las fechas solicitadas
            monthly_predictions_df['year_month'] = monthly_predictions_df['fecha'].dt.to_period('M')
            requested_periods = pd.PeriodIndex([d.to_period('M') for d in requested_dates])
            
            filtered_predictions = monthly_predictions_df[
                monthly_predictions_df['year_month'].isin(requested_periods)
            ].copy()
            
            # CAMBIO 4: Obtener valores reales para el período de predicción (si existen)
            # Buscar datos reales para el período solicitado
            real_data_schema = DemandSearchSchema(
                tipo=2,
                fecha_inicio=search.start_date,
                fecha_fin=search.end_date
            )

            real_data_df = pd.DataFrame()
            try:
                print(f"Consultando datos reales para el período: {search.start_date} - {search.end_date}")
                real_data_raw = demand.get_demands(db, real_data_schema)
                print(f"Respuesta datos reales: {real_data_raw}")
                
                if real_data_raw and real_data_raw.get("lista_fechas") and real_data_raw.get("lista_demandas"):
                    # Los datos vienen en formato "YYYY-MM", los convertimos correctamente
                    real_dates_str = real_data_raw["lista_fechas"]
                    real_values = real_data_raw["lista_demandas"]
                    
                    print(f"Fechas reales: {real_dates_str}")
                    print(f"Valores reales: {real_values}")
                    
                    # Verificar que cada mes esté completo y sea solo de archivos txf
                    validated_dates = []
                    validated_values = []
                    
                    for i, date_str in enumerate(real_dates_str):
                        year_month = pd.to_datetime(date_str).to_period('M')
                        year = year_month.year
                        month = year_month.month
                        
                        # Verificar días completos del mes y que sean solo txf
                        days_in_month = calendar.monthrange(year, month)[1]
                        
                        # Consultar días del mes en la base de datos
                        month_days = db.query(Demand).filter(
                            func.extract('year', Demand.fecha) == year,
                            func.extract('month', Demand.fecha) == month
                        ).all()
                        
                        # Verificar que todos los días estén presentes y sean txf
                        txf_days = [d for d in month_days if d.file_type == 'txf']
                        
                        if len(txf_days) == days_in_month:
                            # Mes completo con solo archivos txf
                            validated_dates.append(date_str)
                            validated_values.append(real_values[i])
                        else:
                            print(f"Mes {year}-{month} descartado: {len(txf_days)} días txf de {days_in_month} totales")
                    
                    if validated_dates:
                        # Crear DataFrame con valores reales validados
                        real_data_df = pd.DataFrame({
                            'year_month_str': validated_dates,
                            'valor_real': validated_values
                        })
                        
                        # Convertir las fechas string a períodos para hacer el merge
                        real_data_df['year_month'] = pd.PeriodIndex(validated_dates, freq='M')
                        print(f"Real data periods validados: {real_data_df['year_month'].tolist()}")
                    else:
                        print("No hay meses completos con archivos txf válidos")
                        
            except Exception as e:
                print(f"No se pudieron obtener valores reales: {str(e)}")
                real_data_df = pd.DataFrame()
            
            # Combinar predicciones con valores reales
            result_df = filtered_predictions.copy()
            print(f"Prediction periods: {filtered_predictions['year_month'].tolist()}")
            
            if not real_data_df.empty:
                print("Haciendo merge de predicciones con datos reales...")
                result_df = result_df.merge(
                    real_data_df[['year_month', 'valor_real']], 
                    on='year_month', 
                    how='left'
                )
                print(f"Resultado después del merge: {result_df[['year_month', 'prediccion', 'valor_real']].head()}")
            else:
                print("No hay datos reales disponibles, asignando None")
                result_df['valor_real'] = None

            # CAMBIO 5: Calcular métricas de error cuando tengamos ambos valores
            result_df['error_porcentual'] = None
            result_df['error_absoluto_mwh'] = None
            
            print(f"Columnas en result_df: {result_df.columns.tolist()}")
            print(f"Shape de result_df: {result_df.shape}")
            
            # Calcular errores solo donde tengamos valores reales
            mask = result_df['valor_real'].notna()
            print(f"Máscara de valores reales disponibles: {mask.sum()} de {len(mask)}")
            
            # Inicializar métricas agregadas
            mape_total = None
            error_minimo_mwh = None
            error_medio_mwh = None
            error_maximo_mwh = None
            error_minimo_porcentual = None
            error_medio_porcentual = None
            error_maximo_porcentual = None
            
            if mask.any():
                print("Calculando errores...")
                # Calcular errores porcentuales
                result_df.loc[mask, 'error_porcentual'] = (
                    (result_df.loc[mask, 'prediccion'] - result_df.loc[mask, 'valor_real']) / 
                    result_df.loc[mask, 'valor_real'] * 100
                )
                
                # Error absoluto convertido a MWh (dividir por 1,000,000)
                result_df.loc[mask, 'error_absoluto_mwh'] = abs(
                    result_df.loc[mask, 'prediccion'] - result_df.loc[mask, 'valor_real']
                ) / 1_000_000  # Convertir de Wh a MWh
                
                # Calcular métricas agregadas para toda la ventana de pronóstico
                try:
                    # MAPE total de la ventana de pronóstico
                    errores_porcentuales_abs = abs(result_df.loc[mask, 'error_porcentual'])
                    mape_total = float(errores_porcentuales_abs.mean())
                    print(f"MAPE total calculado: {mape_total}")
                    
                    # Errores porcentuales para calcular min, medio, max
                    errores_porcentuales = abs(result_df.loc[mask, 'error_porcentual'])
                    error_minimo_porcentual = float(errores_porcentuales.min())
                    error_medio_porcentual = float(errores_porcentuales.mean())
                    error_maximo_porcentual = float(errores_porcentuales.max())
                    print(f"Errores calculados - Min: {error_minimo_mwh}, Medio: {error_medio_mwh}, Max: {error_maximo_mwh}")
                    
                except Exception as e:
                    print(f"Error al calcular métricas agregadas: {str(e)}")
                    mape_total = None
                    error_minimo_mwh = None
                    error_medio_mwh = None
                    error_maximo_mwh = None
            else:
                print("No hay valores reales disponibles para calcular errores")

            # Función auxiliar para convertir valores numpy a tipos nativos de Python
            def convert_to_native(value):
                if pd.isna(value) or value is None:
                    return None
                elif isinstance(value, (np.integer, np.floating)):
                    return float(value)
                elif isinstance(value, np.ndarray):
                    return value.tolist()
                else:
                    return value
            
            print("Preparando respuesta final...")
            
            # Preparar respuesta asegurando que todos los valores sean serializables
            annual_predictions_full = historical_annual_totals + [
                float(pred) for pred in annual_predictions[:num_years_to_predict]
            ]

            response = {
                "annual_predictions": annual_predictions_full,
                "monthly_predictions": {
                    "fechas": result_df['fecha'].dt.strftime('%Y-%m-%d').tolist(),
                    "valores_pronostico": [convert_to_native(val) for val in result_df['prediccion'].tolist()],
                    "valores_reales": [convert_to_native(val) for val in result_df['valor_real'].tolist()],
                    "errores_porcentuales": [convert_to_native(val) for val in result_df['error_porcentual'].tolist()]
                },
                "growth": {str(k): float(v) for k, v in growth.items()},
                "metricas_error": {
                    "mape_total_ventana": mape_total,
                    "error_minimo_%": error_minimo_porcentual,
                    "error_medio_%": error_medio_porcentual,
                    "error_maximo_%": error_maximo_porcentual,
                    "num_meses_con_datos_reales": int(result_df['valor_real'].notna().sum()),
                    "num_meses_totales": int(len(result_df))
                },
                "demand_values_for_table": demand_values_for_table,
                "date_values_for_table": date_values_for_table,
            }
            
            print("Respuesta preparada exitosamente")
            return response
                
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al obtener datos de demanda: {str(e)}")
            return {"message": f"Ocurrió un error en la base de datos: {str(e)}", "status": "error"}
        except ValueError as e:
            print(f"Error de validación: {str(e)}")
            return {"message": str(e), "status": "error"}
        except Exception as e:
            error_msg = str(e)
            print(f"Error en la predicción: {error_msg}")
            print(f"Traceback completo:")
            import traceback
            traceback.print_exc()
            
            # Verificar si el error es "not enough values to unpack"
            if "not enough values to unpack" in error_msg:
                return {
                    "message": "Error en el procesamiento: No hay suficientes datos para realizar la predicción. " +
                            "Verifique que existan variables macroeconómicas para el período solicitado.",
                    "status": "error"
                }
            
            return {"message": f"Ocurrió un error al procesar la solicitud: {error_msg}", "status": "error"}
        
    def get_monthly_percentage(self, db: Session):
        try:
            monthly_demand = db.query(MonthlyDemand).all()
            monthly_data = {}

            for demand in monthly_demand:
                key = f"{demand.year}-{demand.month}"
                monthly_data[key] = {
                    "value": demand.value,
                    "percentage": demand.percentage,
                    "climate_type": demand.climate_type
                }

            return monthly_data
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al crear o actualizar el modelo: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al obtener datos mensuales: {str(e)}")
            raise e
        
    def create_new_model(self, db: Session, search: NewUserModelSchema):
        try:
            model_name = search.model_name
            user_id = search.user_id
            session_id = search.session_id
            start_date = pd.to_datetime(search.start_date).date()
            end_date = pd.to_datetime(search.end_date).date()

            db.add(UsersModels(
                model_name=model_name,
                user_id=user_id,
                session_id=session_id,
                end_date=end_date,
                start_date=start_date
            ))

            db.commit()

            model_id = db.query(UsersModels).filter_by(model_name=model_name).first()
            return {"message": "Modelo creado exitosamente", "model_id": model_id.id}
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al crear nuevo modelo: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al crear nuevo modelo: {str(e)}")
            raise e

    def create_or_update_user_model(self, db: Session, search: CreateOrUpdateUsersModelsSchema):
        try:
            models_name = self.list_user_models(db, search.user_id, search.session_id)
            if search.model_id not in [model["id"] for model in models_name]:
                raise ValueError("El modelo especificado no existe para el usuario y sesión especificados.")
            
            model_id = search.model_id
            dates = search.dates
            values = search.values
            is_model_id = db.query(UsersModelsValues).filter_by(model_id=model_id).first()
            if is_model_id:
                db.query(UsersModelsValues).filter_by(model_id=model_id).delete()
                db.commit()
            for i in range(len(dates)):
                db.add(UsersModelsValues(
                    model_id=model_id,
                    date=pd.to_datetime(dates[i]).date(),
                    value=values[i],
                    climate_type="NORMAL"
                ))

            db.commit()
            return {"message": "Datos del modelo actualizados"}
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al crear o actualizar modelo: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al crear o actualizar modelo: {str(e)}")
            raise e
            
    def list_user_models(self, db: Session, user_id: int, session_id: int):
        try:
            user_models = db.query(UsersModels).filter_by(user_id=user_id, session_id=session_id).all()
            if not user_models:
                raise ValueError("No se encontraron modelos para el usuario especificado.")
            models = []
            for model in user_models:
                models.append({
                    "id": model.id,
                    "model_name": model.model_name,
                    "start_date": model.start_date.strftime('%Y-%m-%d'),
                    "end_date": model.end_date.strftime('%Y-%m-%d'),
                })

            return models
        except SQLAlchemyError as e:	
            db.rollback()	
            print(f"Error al listar modelos de usuario: {str(e)}")	
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al listar modelos de usuario: {str(e)}")
            raise e
        
    def retrieve_user_model_values(self, db: Session, model_id: int):
        try:
            model = db.query(UsersModelsValues).filter_by(model_id=model_id).all()
            if not model:
                raise ValueError("No se encontraron valores para el modelo especificado.")
            

            dates = []
            values = []
            climate_type = []
            for data in model:
                dates.append(data.date.strftime('%Y-%m-%d'))
                values.append(data.value)
                climate_type.append(data.climate_type)
            
            # Get the smallest (earliest) date
            start_date = pd.to_datetime(min(dates))
            
            year_and_value = list(zip(dates, values))
            print("year_and_value", year_and_value)

            year_on_model = list(set([data.date.year for data in model]))
            print("year_on_model", year_on_model)
            

            # Query monthly demand data up to the start date year
            # Query monthly demand data up to the start date (using both year and month)
            historic_demands = db.query(MonthlyDemand.value, MonthlyDemand.year, MonthlyDemand.month)\
                .filter(
                    # Exclude data from after start date (including specific month)
                    ((MonthlyDemand.year < start_date.year) |
                     ((MonthlyDemand.year == start_date.year) & (MonthlyDemand.month < start_date.month)))
                )\
                .order_by(MonthlyDemand.year, MonthlyDemand.month)\
                .all()
                
            # Create separate arrays for demands and dates while preserving their order
            demand_values_for_table = []
            date_values_for_table = []

            for record in historic_demands:
                demand_values_for_table.append(record.value)
                # Format date as 'YYYY-MM-01'
                date_values_for_table.append(f"{record.year}-{record.month:02d}-01")

            # Calcular lo valores anuales del modelo almacenado
            annual_model_values = {}
            for year in year_on_model:
                annual_model_values[year] = sum([data.value for data in model if data.date.year == year])

            annual_totals = {str(year): total for year, total in annual_model_values.items()}

            historic_demand_values = db.query(YearlyDemand).all()
            historic_demand_values = {
               str(demand.year): demand.demand for demand in historic_demand_values if demand.year not in year_on_model
            }
            print("historic_demand_values", historic_demand_values)
            # Agregar el valor anual del modelo a los valores históricos
            for year in year_on_model:
                historic_demand_values[str(year)] = annual_model_values[year]
            print("historic_demand_values", historic_demand_values)
            # Realizar el array de crecimiento 
            growth = {}
            years = list(historic_demand_values.keys())
            for i in range(1, len(years)):
                growth[years[i]] = (historic_demand_values[years[i]] - historic_demand_values[years[i-1]]) / historic_demand_values[years[i-1]]

            years_sorted = sorted(historic_demand_values.keys(), key=lambda y: int(y))
            annual_predictions = [float(historic_demand_values[str(year)]) for year in years_sorted]
            
            return {
                "dates": dates,
                "values": values,
                "climate_type": climate_type,
                "growth": growth,
                "annual_totals": annual_totals,	
                "annual_predictions": annual_predictions,
                "demand_values_for_table": demand_values_for_table,
                "date_values_for_table": date_values_for_table
            }
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al obtener valores de modelo: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al obtener valores de modelo: {str(e)}")
            raise e
        
    def get_monthly_info(self, db: Session, search: MonthlyInfoSchema):
        try:
            year = search.year
            month = search.month
            prediction = search.predicition

            historic_values = db.query(MonthlyDemand.value).filter_by(year=year, month=month).all()

            number_of_working_days, number_of_saturdays, number_of_sundays_holidays = self.get_number_of_days(year, month)        
            results = db.query(
                MonthlyDemand.year,
                func.max(MonthlyDemand.value).label('max_value'),
                func.min(MonthlyDemand.value).label('min_value'),
                func.avg(MonthlyDemand.value).label('avg_value')
            ).group_by(MonthlyDemand.year).filter_by(year=year).all()

            if not results:
                raise ValueError("No se encontraron datos para el ano y mes especificados.")
            
            
            return {
                "number_of_working_days": number_of_working_days,
                "number_of_saturdays": number_of_saturdays,
                "number_of_sundays_holidays": number_of_sundays_holidays,
                "max_value": round(results[0].max_value, 2),
                "min_value": round(results[0].min_value, 2),
                "avg_value": round(results[0].avg_value, 2),
                "error": round(abs(prediction - historic_values[0][0]), 2) / historic_values[0][0] if historic_values else None
            }
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al obtener información mensual: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al obtener información mensual: {str(e)}")
            raise e
        
#perfiles de demanda
    def get_day_behavior(self, search: DayBehaviorSchema, db: Session):
        # Validaciones iniciales para los nuevos campos
        if search.model_type not in [0, 1]:
            raise ValueError("model_type debe ser 0 o 1")
        if search.user_id is None or search.session_id is None:
            raise ValueError("user_id y session_id son requeridos para obtener el modelo")
        
        # Determinar las fechas a usar según el tipo de modelo
        if search.model_type == 0:
            # Modelo guardado en DB: obtener fechas del modelo almacenado
            try:
                # Buscar el modelo en la lista de modelos del usuario
                models = self.list_user_models(db, user_id=search.user_id, session_id=search.session_id)  # Ajustar según contexto
                target_model = None
                for model in models:
                    if model["id"] == search.model_id:
                        target_model = model
                        break
                
                if not target_model:
                    raise ValueError(f"No se encontró el modelo con ID {search.model_id}")
                
                # Usar las fechas del modelo guardado
                start_date = pd.to_datetime(target_model["start_date"])
                end_date = pd.to_datetime(target_model["end_date"])
                
                # Obtener los valores del pronóstico del modelo guardado
                model_values = self.retrieve_user_model_values(db, search.model_id)
                forecast_values_dict = {}
                for i, date_str in enumerate(model_values["dates"]):
                    date_obj = pd.to_datetime(date_str)
                    year_month = f"{date_obj.year}-{date_obj.month:02d}"
                    # Convertir de Wh a MWh si es necesario
                    value_mwh = model_values["values"][i]
                    value_mwh = value_mwh / 1000000

                    forecast_values_dict[year_month] = value_mwh
                
            except Exception as e:
                raise ValueError(f"Error al obtener modelo guardado: {str(e)}")
                
        else:
            # Modelo básico (1, 2, 3): usar fechas enviadas en el endpoint
            start_date = pd.to_datetime(search.start_date)
            end_date = pd.to_datetime(search.end_date)
            
            # Generar pronóstico usando los modelos básicos
            try:
                from ..config.schemas import ForecastSchema
                forecast_search = ForecastSchema(
                    id=search.model_id,  # 1=RF, 2=XGBoost, 3=SARIMAX
                    user_id=search.user_id,  # Ajustar según contexto
                    session_id=search.session_id,  # Ajustar según contexto
                    start_date=search.start_date,
                    end_date=search.end_date
                )
                
                forecast_result = self.predict(forecast_search, db)
                
                # Convertir resultado a diccionario por mes
                forecast_values_dict = {}
                
                # CORRECIÓN: Acceso correcto a la estructura de predict()
                if "monthly_predictions" in forecast_result:
                    fechas = forecast_result["monthly_predictions"]["fechas"]
                    valores = forecast_result["monthly_predictions"]["valores_pronostico"]
                    
                    for i, date_str in enumerate(fechas):
                        date_obj = pd.to_datetime(date_str)
                        year_month = f"{date_obj.year}-{date_obj.month:02d}"
                        # Los valores del predict ya deberían estar en MWh
                        value_mwh = valores[i]
                        value_mwh = value_mwh / 1000000
                        forecast_values_dict[year_month] = value_mwh
                else:
                    raise ValueError("La respuesta del modelo no contiene 'monthly_predictions'")
                    
            except Exception as e:
                raise ValueError(f"Error al generar pronóstico con modelo básico: {str(e)}")

        start_year = start_date.year
        end_year = end_date.year    

        if start_date > end_date:
            raise ValueError("La fecha de inicio no puede ser mayor a la fecha final")
        if search.lag < 0:
            raise ValueError("El lag debe ser un número positivo")
        
        lag = search.lag

        # Generar fechas con lag (IGUAL que función original)
        dates = []
        for i in range(lag+1):
            dates.append(((start_date - pd.DateOffset(years=i)).strftime('%Y-%m-%d'), 
                        (end_date - pd.DateOffset(years=i)).strftime('%Y-%m-%d')))

        demands = []

        for date in dates:
            start_date_query = pd.to_datetime(date[0]).date()
            end_date_query = pd.to_datetime(date[1]).date()
            demand = db.query(Demand).filter(
                Demand.fecha >= start_date_query, 
                Demand.fecha <= end_date_query,
                Demand.file_type == 'txf' 
            ).all()
            demands.extend(demand)
        
        years = [demand.fecha.year for demand in demands]
        years = list(set(years))

        monthly_demands = db.query(MonthlyDemand.month, MonthlyDemand.year, MonthlyDemand.value).filter(MonthlyDemand.year.in_(years)).all()
        
        # IGUAL que función original - crear DataFrame con todas las horas
        df_demands = pd.DataFrame([{
            'fecha': demand.fecha,
            'ano': demand.fecha.year,
            'mes': demand.fecha.month,
            'tipo_fecha': demand.tipo_fecha,
            'total': demand.total,
            **{f'hora_{i}': getattr(demand, f'hora_{i}') for i in range(1, 25)}
        } for demand in demands])

        monthly_demands_df = pd.DataFrame([{
            'ano': demand.year,
            'mes': demand.month,
            'value': demand.value
        } for demand in monthly_demands if (demand.year, demand.month) in [(demand.fecha.year, demand.fecha.month) for demand in demands]])

        # IGUAL que función original - merge
        df = df_demands.merge(monthly_demands_df, on=['ano', 'mes'], how='left')

        # IGUAL que función original - normalizar para crear perfiles
        for i in range(1, 25):
            df[f'hora_{i}'] = df[f'hora_{i}'] / df['value']
            
        # IGUAL que función original - agrupar para obtener perfiles promedio
        perfiles = df.groupby(['mes', 'tipo_fecha']).agg({
            **{f'hora_{i}': 'mean' for i in range(1, 25)}
        }).reset_index()
        
        # IGUAL que función original - promedio mensual
        df_monthly_demand_average = monthly_demands_df.groupby(['mes']).agg({
            'value': 'mean'
        }).reset_index()

        result = []
        
        # CAMBIO PRINCIPAL: En lugar de iterar solo meses, iterar combinaciones ano-mes del pronóstico
        forecast_periods = []
        for key in forecast_values_dict.keys():
            year, month = key.split('-')
            forecast_periods.append((int(year), int(month)))
        
        # Ordenar por ano y mes
        forecast_periods = sorted(forecast_periods)
        
        for year, mes in forecast_periods:
            # Obtener número de días por tipo - USAR EL AÑO CORRECTO
            number_of_working_days, number_of_saturdays, number_of_sundays_holidays = self.get_number_of_days(year, mes)
            
            # Obtener demanda pronosticada para este mes
            year_month_key = f"{year}-{mes:02d}"
            forecast_value_mwh = forecast_values_dict.get(year_month_key, 0)
            forecast_value_gwh = forecast_value_mwh / 1000  
            
            # CAMBIO: Obtener demanda histórica promedio para referencia
            historical_value = df_monthly_demand_average[df_monthly_demand_average['mes'] == mes]['value'].values
            historical_value_mwh = historical_value[0] / 1000000 if len(historical_value) > 0 and historical_value[0] > 1000000 else (historical_value[0] / 1000 if len(historical_value) > 0 and historical_value[0] > 1000 else (historical_value[0] if len(historical_value) > 0 else 0))
            
            month_data = {
                'mes': int(mes),
                'year': int(year),  # NUEVO: Agregar el ano
                'tipos': [],
                'dias_laborales': number_of_working_days,
                'sabados': number_of_saturdays,
                'domingos_festivos': number_of_sundays_holidays,
                'demanda_pronosticada_gwh': round(forecast_value_gwh, 2)  # CAMBIO: era 'demanda_promedio'
            }
            
            # Buscar perfiles para este mes (pueden estar en cualquier ano del histórico)
            perfiles_mes = perfiles[perfiles['mes'] == mes]
            
            if perfiles_mes.empty:
                # Si no hay perfiles para este mes, crear vacíos o usar promedio general
                # Por ahora mantenemos vacío para ser consistente
                result.append(month_data)
                continue
                
            # IGUAL que función original - procesar cada tipo
            for tipo in perfiles_mes['tipo_fecha'].unique():
                perfil_data = perfiles_mes[perfiles_mes['tipo_fecha'] == tipo]
                
                if not perfil_data.empty:
                    # CAMBIO CLAVE: En lugar de usar demanda histórica, usar pronóstico
                    horas = {}
                    for i in range(1, 25):
                        perfil_normalizado = perfil_data[f'hora_{i}'].values[0]  # Proporción (0-1)
                        # CAMBIO: Multiplicar por pronóstico en lugar de histórico
                        valor_absoluto_mwh = perfil_normalizado * forecast_value_mwh
                        horas[f'hora_{i}'] = round(valor_absoluto_mwh, 6)
                    
                    tipo_data = {
                        'tipo_fecha': int(tipo),  # IGUAL que original
                        'horas': horas  # CAMBIO: Valores ajustados por pronóstico
                    }
                    
                    month_data['tipos'].append(tipo_data)
            
            result.append(month_data)

        # Agregar información adicional del pronóstico de referencia (NUEVO)
        forecast_reference = {
            'model_type': search.model_type,
            'model_id': search.model_id,
            'period': f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
            'description': f"{'Modelo guardado' if search.model_type == 0 else 'Modelo básico'} ID: {search.model_id}",
            'forecast_dates': list(forecast_values_dict.keys()),
            'forecast_values_gwh': [round(v, 2) for v in forecast_values_dict.values()],
            'total_forecast_gwh': round(sum(forecast_values_dict.values()), 2),
            'method_used': 'Potencia pronosticada'  # CAMBIO: antes era "potencia promedio"
        }

        return {
            'data': result,
            'summary': {
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'lag_years': lag,
                'total_months_analyzed': len(result),
                'forecast_reference': forecast_reference  # NUEVO
            }
        }

    def get_number_of_days(self, year, month):
            """
            Obtiene el número de días laborales, sábados y domingos/festivos en un mes
            """
            import calendar
            from datetime import datetime, timedelta
            from holidays_co import get_colombia_holidays_by_year
            
            # Obtener días festivos de Colombia usando el módulo del proyecto
            colombia_holidays = get_colombia_holidays_by_year(year)
            # Convertir fechas string a objetos date si es necesario
            holiday_dates = []
            for holiday in colombia_holidays:
                if isinstance(holiday, str):
                    holiday_dates.append(datetime.strptime(holiday, '%Y-%m-%d').date())
                else:
                    holiday_dates.append(holiday)
            
            # Obtener el primer y último día del mes
            first_day = datetime(year, month, 1).date()
            last_day = datetime(year, month, calendar.monthrange(year, month)[1]).date()
            
            working_days = 0
            saturdays = 0
            sundays_holidays = 0
            
            current_date = first_day
            while current_date <= last_day:
                if current_date in holiday_dates or current_date.weekday() == 6:  # Domingo o festivo
                    sundays_holidays += 1
                elif current_date.weekday() == 5:  # Sábado
                    saturdays += 1
                else:  # Día laboral
                    working_days += 1
                current_date += timedelta(days=1)
            
            return working_days, saturdays, sundays_holidays


    def create_type_year_list_service(self, search: CreateTypeYearListSchema, db: Session):
        try:
            user_id = search.user_id
            session_id = search.session_id
            years = db.query(YearlyDemand.year).all()
            years = [year[0] for year in years]
            for year in years:
                db.add(TypeYear(
                    year=year,
                    type="TIPICO",
                    user_id=user_id,
                    session_id=session_id
                ))
            db.commit()
            return {"message": "Lista de anos creada exitosamente"}
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al crear lista de anos: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al crear lista de anos: {str(e)}")
            raise e

    def update_type_year_list_service(self, search: UpdateTypeYearListSchema, db: Session):
        try:
            user_id = search.user_id
            session_id = search.session_id
            years = search.year
            types = search.type
            for i in range(len(years)):
                db.query(TypeYear).filter_by(year=years[i], user_id=user_id, session_id=session_id).update({
                    "type": types[i]
                })
            db.commit()
            return {"message": "Lista de anos actualizada exitosamente"}
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al actualizar lista de anos: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al actualizar lista de anos: {str(e)}")
            raise e
        
    def get_type_year_list_service(self, user_id: int, session_id: int, db: Session):
        try:
            years = db.query(TypeYear).filter_by(user_id=user_id, session_id=session_id).all()
            if not years:
                raise ValueError("No se encontraron anos para el usuario especificado.")
            years_list = []
            for year in years:
                years_list.append({
                    "year": year.year,
                    "type": year.type
                })
            # Sort the list by year in ascending order
            years_list = sorted(years_list, key=lambda x: x["year"])
            return years_list
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al obtener lista de anos: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al obtener lista de anos: {str(e)}")
            raise e
        
    def procesar_datos_y_crear_excel(self, mothly_demands, predictions):
        wb = Workbook()
        ws = wb.active
        ws.title = "CARIBE_MAR"
        
        # Configurar estilos
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Crear estructura base (headers)
        ws['A1'] = "Mercado Regulado"
        ws.merge_cells('A1:Q2')
        ws['T1'] = "Crecimiento"
        ws.merge_cells('T1:AH1')
        ws['T2'] = "Mensual"
        ws.merge_cells('T2:AE2')
        ws['AF2'] = "Anual"
        ws['AG2'] = "Semestral"
        ws.merge_cells('AG2:AH2')
        
        # Headers tercera fila
        headers_row_3 = [
            "CM", "Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", 
            "Oct", "Nov", "Dic", "Semestre 1", "Semestre 2", "Total Año", "% de Crecimiento",
            "", "Año", "Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep",
            "Oct", "Nov", "Dic", "Total", "Semestre 1", "Semestre 2"
        ]
        
        for col, header in enumerate(headers_row_3, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
        
        # Procesar datos históricos y predicciones
        all_data = {}
        
        # Procesar datos históricos y MWh
        for demand in mothly_demands:
            year = demand['year']
            month = demand['month']
            value = demand['value'] / 1_000_000
            
            if year not in all_data:
                all_data[year] = {'months': [None]*12}
            all_data[year]['months'][month-1] = value
        
        # Procesar predicciones y MWh
        for fecha, valor in zip(predictions['monthly_predictions']['fechas'], 
                            predictions['monthly_predictions']['valores_pronostico']):
            date = datetime.strptime(fecha, '%Y-%m-%d')
            year = date.year
            month = date.month
            value = valor / 1_000_000
            
            if year not in all_data:
                all_data[year] = {'months': [None]*12}
            all_data[year]['months'][month-1] = value
        
        # Llenar el Excel con los datos y fórmulas
        current_row = 4
        for year in sorted(all_data.keys()):
            # Columna ano con formato de número entero (columna A - Mercado Regulado)
            year_cell = ws.cell(row=current_row, column=1, value=year)
            year_cell.number_format = '0'
            
            # Valores mensuales
            for month in range(12):
                value = all_data[year]['months'][month]
                if value is not None:
                    cell = ws.cell(row=current_row, column=month+2, value=value)
                    cell.number_format = '#,##0.00'  # Mantener decimales para los valores en GW
            
            # Fórmulas para semestres y total ano
            sem1_col = get_column_letter(14)  # Columna N
            sem2_col = get_column_letter(15)  # Columna O
            total_col = get_column_letter(16) # Columna P
            growth_col = get_column_letter(17) # Columna Q - % de Crecimiento
            
            # Semestre 1 (suma de Ene a Jun)
            ws[f'{sem1_col}{current_row}'] = f'=SUM(B{current_row}:G{current_row})'
            
            # Semestre 2 (suma de Jul a Dic)
            ws[f'{sem2_col}{current_row}'] = f'=SUM(H{current_row}:M{current_row})'
            
            # Total ano
            ws[f'{total_col}{current_row}'] = f'=SUM({sem1_col}{current_row}:{sem2_col}{current_row})'
            
            # ===== CORRECCIÓN 1: % de Crecimiento Anual (columna Q) =====
            if current_row > 4:  # Si hay ano anterior
                ws[f'{growth_col}{current_row}'] = f'=IF({total_col}{current_row-1}=0,0,({total_col}{current_row}/{total_col}{current_row-1}-1)*100)'
                # Aplicar formato de porcentaje
                ws[f'{growth_col}{current_row}'].number_format = '#,##0.00"%"'
            else:
                # Para el primer ano, no hay crecimiento
                ws[f'{growth_col}{current_row}'] = 0
                ws[f'{growth_col}{current_row}'].number_format = '#,##0.00"%"'
            

            year_growth_col = get_column_letter(19)  # Columna S - Año en sección Crecimiento
            year_growth_cell = ws.cell(row=current_row, column=19, value=year)
            year_growth_cell.number_format = '0'
            
            # Cálculos de crecimiento mensual (respecto al ano anterior)
            if current_row > 4:  # Si hay ano anterior
                # Crecimientos mensuales (columnas T a AE)
                for month in range(12):
                    growth_monthly_col = get_column_letter(20 + month)  # Columnas T-AE
                    value_col = get_column_letter(2 + month)  # Columnas B-M
                    ws[f'{growth_monthly_col}{current_row}'] = f'=IF({value_col}{current_row-1}=0,0,({value_col}{current_row}/{value_col}{current_row-1}-1)*100)'
                    # Aplicar formato de porcentaje
                    ws[f'{growth_monthly_col}{current_row}'].number_format = '#,##0.00"%"'
                
                # Total crecimiento anual (columna AF)
                ws[f'AF{current_row}'] = f'=IF({total_col}{current_row-1}=0,0,({total_col}{current_row}/{total_col}{current_row-1}-1)*100)'
                ws[f'AF{current_row}'].number_format = '#,##0.00"%"'
                
                # Crecimiento semestral (columnas AG y AH)
                ws[f'AG{current_row}'] = f'=IF({sem1_col}{current_row-1}=0,0,({sem1_col}{current_row}/{sem1_col}{current_row-1}-1)*100)'
                ws[f'AG{current_row}'].number_format = '#,##0.00"%"'
                
                ws[f'AH{current_row}'] = f'=IF({sem2_col}{current_row-1}=0,0,({sem2_col}{current_row}/{sem2_col}{current_row-1}-1)*100)'
                ws[f'AH{current_row}'].number_format = '#,##0.00"%"'
            else:
                # Para el primer ano, todos los crecimientos son 0
                for month in range(12):
                    growth_monthly_col = get_column_letter(20 + month)
                    ws[f'{growth_monthly_col}{current_row}'] = 0
                    ws[f'{growth_monthly_col}{current_row}'].number_format = '#,##0.00"%"'
                
                # Crecimientos anuales y semestrales = 0
                ws[f'AF{current_row}'] = 0
                ws[f'AF{current_row}'].number_format = '#,##0.00"%"'
                ws[f'AG{current_row}'] = 0  
                ws[f'AG{current_row}'].number_format = '#,##0.00"%"'
                ws[f'AH{current_row}'] = 0
                ws[f'AH{current_row}'].number_format = '#,##0.00"%"'
            
            current_row += 1
        
        # Ajustar formato
        for row in range(1, current_row):
            for col in range(1, 35):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = center_alignment
                
                # Aplicar formato numérico según el tipo de columna
                if isinstance(cell.value, (int, float)):
                    if col == 1 or col == 19:  # Columnas de anos (A y S)
                        cell.number_format = '0'
                    elif col == 17 or col >= 20:  # Columnas de porcentajes (Q y T en adelante)
                        if not cell.number_format or cell.number_format == 'General':
                            cell.number_format = '#,##0.00"%"'
                    else:  # Otras columnas numéricas (valores en GW)
                        if not cell.number_format or cell.number_format == 'General':
                            cell.number_format = '#,##0.00'
        
        # Ajustar anchos de columna para mejor visualización
        for col in range(1, 35):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        return wb


    def predict_excel(self, search: ForecastSchema, db: Session):
        try:
            # Obtener datos históricos
            mothly_demands = db.query(MonthlyDemand.year, MonthlyDemand.month, MonthlyDemand.value).all()
            mothly_demands = [{
                'year': demand.year,
                'month': demand.month,
                'value': demand.value
            } for demand in mothly_demands]
            
            # Obtener predicciones
            predictions = self.predict(search, db)
            
            # Crear el Excel
            excel_workbook = self.procesar_datos_y_crear_excel(mothly_demands, predictions)
            
            # Guardar el Excel en un buffer de memoria
            excel_buffer = io.BytesIO()
            excel_workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error processing Excel: {str(e)}")
        
    def create_day_behavior_excel(self, day_behavior_data, end_year):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Comportamiento_Diario"
        
        # Configurar estilos
        header_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Headers (IGUAL que función original)
        headers = ["AÑO", "MES", "Dias\\Horas"]
        headers.extend([f"H{str(i).zfill(2)}" for i in range(1, 25)])  # H01 a H24
        headers.append("#dias")
        
        try:
            # Aplicar headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border
            
            # Llenar datos
            current_row = 2
            # Mapeo correcto de tipos (IGUAL que función original)
            tipos_fecha = {
                0: "LABORAL",
                1: "SABADO",
                2: "DOMINGO/FESTIVO"
            }
            
            # CAMBIO: Acceder a day_behavior_data['data'] en lugar de day_behavior_data directamente
            data_list = day_behavior_data['data'] if isinstance(day_behavior_data, dict) else day_behavior_data
            
            for month_data in data_list:
                mes = month_data['mes']
                
                for tipo in month_data['tipos']:
                    tipo_fecha = tipo['tipo_fecha']
                    
                    # Año, Mes y Tipo de día (IGUAL que función original)
                    ws.cell(row=current_row, column=1, value=end_year)
                    ws.cell(row=current_row, column=2, value=mes)
                    ws.cell(row=current_row, column=3, value=tipos_fecha[tipo_fecha])
                    
                    # Valores por hora (IGUAL que función original)
                    for hora in range(1, 25):
                        value = tipo['horas'][f'hora_{hora}']
                        ws.cell(row=current_row, column=hora + 3, value=value)
                    
                    # Número de días según el tipo correcto (IGUAL que función original)
                    if tipo_fecha == 0:
                        dias = month_data['dias_laborales']
                    elif tipo_fecha == 1:
                        dias = month_data['sabados']
                    else:  # tipo_fecha == 2
                        dias = month_data['domingos_festivos']
                    
                    ws.cell(row=current_row, column=28, value=dias)
                    
                    # Aplicar formato (IGUAL que función original)
                    for col in range(1, 29):
                        cell = ws.cell(row=current_row, column=col)
                        cell.border = border
                        cell.alignment = center_alignment
                        if col > 3 and col < 28:  # Columnas de horas
                            # CAMBIO: Actualizar formato para valores en MWh
                            cell.number_format = '#,##0.000000'  # Mejor formato para MWh
                    
                    current_row += 1
            
            # NUEVO: Agregar hoja adicional con información del pronóstico de referencia
            if isinstance(day_behavior_data, dict) and 'summary' in day_behavior_data:
                ws_summary = wb.create_sheet("Resumen_Pronostico")
                
                # Headers de resumen
                ws_summary.cell(row=1, column=1, value="Información del Pronóstico")
                ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
                
                summary = day_behavior_data['summary']
                forecast_ref = summary.get('forecast_reference', {})
                
                # Información básica
                info_rows = [
                    ("Período:", f"{summary.get('start_date', '')} to {summary.get('end_date', '')}"),
                    ("Lag (anos):", summary.get('lag_years', 0)),
                    ("Tipo de modelo:", forecast_ref.get('model_type', 'N/A')),
                    ("ID del modelo:", forecast_ref.get('model_id', 'N/A')),
                    ("Descripción:", forecast_ref.get('description', 'N/A')),
                    ("Método usado:", forecast_ref.get('method_used', 'N/A')),
                    ("Total pronóstico (MWh):", forecast_ref.get('total_forecast_mwh', 0))
                ]
                
                row = 3
                for label, value in info_rows:
                    ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
                    ws_summary.cell(row=row, column=2, value=value)
                    row += 1
                
                # Detalle mensual del pronóstico
                if 'forecast_dates' in forecast_ref and 'forecast_values_mwh' in forecast_ref:
                    row += 2
                    ws_summary.cell(row=row, column=1, value="Pronóstico Mensual (MWh)").font = Font(bold=True)
                    row += 1
                    
                    ws_summary.cell(row=row, column=1, value="Mes").font = Font(bold=True)
                    ws_summary.cell(row=row, column=2, value="Demanda (MWh)").font = Font(bold=True)
                    row += 1
                    
                    for i, (date, value) in enumerate(zip(forecast_ref['forecast_dates'], forecast_ref['forecast_values_mwh'])):
                        ws_summary.cell(row=row, column=1, value=date)
                        ws_summary.cell(row=row, column=2, value=value)
                        ws_summary.cell(row=row, column=2).number_format = '#,##0.00'
                        row += 1
                
                # Ajustar anchos de columna para la hoja de resumen
                ws_summary.column_dimensions['A'].width = 25
                ws_summary.column_dimensions['B'].width = 20
            
            # Ajustar anchos de columna (IGUAL que función original)
            for col in range(1, 29):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 12
            
            ws.column_dimensions['C'].width = 15
            
            return wb
        except Exception as e:
            print(f"Error en create_day_behavior_excel: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            raise

    def day_behavior_excel(self, search: DayBehaviorSchema, db: Session):
        try:
            day_behavior_data = self.get_day_behavior(search, db)
            if not day_behavior_data:
                raise ValueError("No se encontraron datos para el período especificado")
                
            print("Creating Excel workbook...")
            end_date = pd.to_datetime(search.end_date)
            excel_workbook = self.create_day_behavior_excel(day_behavior_data, end_date.year)
            
            print("Saving to buffer...")
            excel_buffer = io.BytesIO()
            excel_workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
        except Exception as e:
            print(f"Error completo: {str(e)}")
            print(f"Tipo de error: {type(e)}")
            import traceback
            print(f"Traceback: {traceback.format_exc()}")
            raise HTTPException(
                status_code=500, 
                detail=f"Error processing Excel: {str(e)}\nTraceback: {traceback.format_exc()}"
            )
        
    def update_monthly_type(self, search: UpdateMonthlyTypeSchema, db: Session):
        try:
            monthly_json = search.values 
            for year in monthly_json:
                for month in monthly_json[year]:
                    db.query(MonthlyDemand).filter_by(year=year, month=month).update({
                        "climate_type": monthly_json[year][month]
                    })

            db.commit()
            return {"message": "Datos actualizados exitosamente"}
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al actualizar datos mensuales: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al actualizar datos mensuales: {str(e)}")
            raise e
        

    def get_last_year_on_db(self, db: Session):
        try:
            demand = DemandService()
            year_range = demand.get_first_and_last_date(db)
            last_date = year_range["last_date"].strftime('%Y-%m-%d')
            last_year = int(pd.to_datetime(last_date).year)
            return {
                "last_year": last_year
            }
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al obtener último ano: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al obtener último ano: {str(e)}")
            raise e

    def list_historic_years(self, db: Session):
        try:
            # Extract all years from December 31st dates
            years = db.query(func.extract('year', Demand.fecha).label('year'))\
                     .filter(func.extract('month', Demand.fecha) == 12,
                             func.extract('day', Demand.fecha) == 31)\
                     .distinct()\
                     .order_by('year')\
                     .all()
            
            return sorted([int(year[0]) for year in years])
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al obtener anos históricos: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al obtener anos históricos: {str(e)}")
            raise e

    def get_monthly_demand_by_year(self, year: int, db: Session):
        try:
            monthly_demand = db.query(MonthlyDemand).filter_by(year=year).all()
            if not monthly_demand:
                raise ValueError("No se encontraron datos para el ano especificado.")
            
            monthly_data = {}
            monthly_data = {}
            
            # Create entries with YYYY-MM format
            for demand in monthly_demand:
                month_str = f"0{demand.month}" if demand.month < 10 else str(demand.month)
                key = f"{year}-{month_str}"
                monthly_data[key] = {
                    "value": demand.value,
                    "percentage": demand.percentage,
                    "climate_type": demand.climate_type
                }
            
            # Sort the dictionary by keys (which are now in YYYY-MM format)
            sorted_data = {k: monthly_data[k] for k in sorted(monthly_data.keys())}
            
            return sorted_data
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al obtener demanda mensual por ano: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al obtener demanda mensual por ano: {str(e)}")
            raise e
        
    def change_model_monthly_type(self, search: ForecastTypeMonthSchema, db: Session):
        try:
            model_id = search.model_id
            dates = search.dates
            types = search.types
            model = db.query(UsersModelsValues).filter_by(model_id=model_id).first()
            if not model:
                raise ValueError("El modelo especificado no existe.")
            
            def change_type_month_value(value, current_type, new_type):
                if current_type == "NORMAL":
                    if new_type == "NIÑA":
                        return value * 0.9946
                    elif new_type == "NIÑO":
                        return value * 1.0104
                    else:
                        return value
                elif current_type == "NIÑA":
                    if new_type == "NORMAL":
                        return value / 0.9946
                    elif new_type == "NIÑO":
                        return ( value / 1.0104 ) * 0.9946
                    else:
                        return value
                elif current_type == "NIÑO":
                    if new_type == "NORMAL":
                        return value / 1.0104
                    elif new_type == "NIÑA":
                        return ( value / 0.9946 ) * 1.0104
                    else:
                        return value
            
            model_user = db.query(UsersModelsValues).filter_by(model_id=model_id).all()
            model_dates = [model.date.strftime('%Y-%m-%d') for model in model_user]
            model_values = [model.value for model in model_user]
            model_types = [model.climate_type for model in model_user]
            list_model = zip(model_dates, model_values, model_types)
            for item in list_model:
                date = item[0]
                value = item[1]
                current_type = item[2]
                if date in dates:
                    index = dates.index(date)
                    new_type = types[index]
                    new_value = change_type_month_value(value, current_type, new_type)
                    db.query(UsersModelsValues).filter_by(model_id=model_id, date=pd.to_datetime(date).date()).update({
                        "value": new_value,
                        "climate_type": new_type
                    })
            db.commit()

            return self.retrieve_user_model_values(model_id=model_id, db=db)
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al cambiar tipo de modelo mensual: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al cambiar tipo de modelo mensual: {str(e)}")
            raise e
    
    def change_model_based_on_year(self, search: ForecastModelSaveBasedOnYear, db: Session):
        try:
            model_id = search.model_id
            year = search.year
            predict_year = search.predict_year
            model_values = db.query(UsersModelsValues).filter_by(model_id=model_id).all()
            if not model_values:
                raise ValueError("El modelo especificado no existe.")
            year_info = self.get_monthly_demand_by_year(year, db)

            model_values_dates = [model.date for model in model_values if str(predict_year) in str(model.date)]
            if not model_values_dates:
                raise ValueError("No se encontraron fechas para el modelo especificado.")
            months_to_change = [model.date.month for model in model_values if str(predict_year) in str(model.date)]
            model_values = [model.value for model in model_values if str(predict_year) in str(model.date)]

            if len(months_to_change) != 12:
                partial_total = 0
                for month in months_to_change:
                    partial_total += year_info[f"{year}-{str(month).zfill(2)}"]["value"]
                partial_percentage_dict = {}
                for month in months_to_change:
                    partial_percentage_dict[month] = year_info[f"{year}-{str(month).zfill(2)}"]["value"] / partial_total

                percentages = partial_percentage_dict
            else:
                percentages = {}
                for month in months_to_change:
                    percentages[month] = year_info[f"{year}-{str(month).zfill(2)}"]["percentage"]

            total_model = sum(model_values)
            model_to_change = zip(model_values_dates, model_values, months_to_change)
            for month in months_to_change:
                date = model_values_dates[months_to_change.index(month)]
                if isinstance(date, pd.Timestamp):
                    date = date.date()
                elif isinstance(date, datetime):
                    date = date.date()
                # Si ya es date, no pasa nada
                value = model_values[months_to_change.index(month)]
                percentage = percentages[month]
                new_value = total_model * percentage
                db.query(UsersModelsValues).filter_by(model_id=model_id, date=date).update({
                    "value": new_value,
                    "climate_type": year_info[f"{year}-{str(month).zfill(2)}"]["climate_type"]
                })
            db.commit()
            return self.retrieve_user_model_values(model_id=model_id, db=db)
        except SQLAlchemyError as e:
            db.rollback()
            print(f"Error al cambiar modelo basado en ano: {str(e)}")
            return {"message": "Ocurrió un error al procesar la solicitud.", "status": "error"}
        except Exception as e:
            print(f"Error al cambiar modelo basado en ano: {str(e)}")
            raise e
            




            