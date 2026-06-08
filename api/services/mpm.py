from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError
import os
from ..other_models.demand import LastDemandDocument, Demand
from ..other_models.climate import Climate
from sqlalchemy import func, case
from datetime import datetime, timedelta
from .functions.xgb_only import DemandOnlyPredictor
from .functions.xgb_temp import DemandPredictor
from .functions.xgb_varias import DemandVariasPredictor
from .functions.xgBoost_multimodelo import DemandPredictorMultiModelo
from .functions.xgBoost_varias_multimodelo import DemandVariasMultiModelo
from api.config.schemas import MPMPredictSchema, SearchDayClimateSchema, SearchAllClimateTypeSchema
import joblib
import pandas as pd
import numpy as np
import calendar
from .analysis import ClimateService    
from holidays_co import get_colombia_holidays_by_year
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
from typing import Optional

class MPMService:
    def _parse_file_date(self, year: str, filename: str) -> tuple[str, datetime]:
        """Parse file date from filename and return formatted date string and datetime object"""
        month = filename[4:6]
        day = filename[6:8]
        date_str = f"{year}-{month}-{day}"
        return date_str, datetime.strptime(date_str, "%Y-%m-%d")

    def update_last_demand_documents_from_filesystem(self, db: Session, base_path: Optional[str] = None):
        """
        Actualiza la tabla LastDemandDocument basandose en archivos ADEM disponibles en disco.
        """
        try:
            base_path = base_path or os.getenv("ADEM_ROUTE") or "files/adem/"
            latest_by_type = {}

            for root, _, files in os.walk(base_path):
                for filename in files:
                    filename_lower = filename.lower()
                    if not filename_lower.startswith("adem"):
                        continue
                    if not (filename_lower.endswith(".txf") or filename_lower.endswith(".txr") or filename_lower.endswith(".tx2")):
                        continue

                    path_parts = os.path.normpath(root).split(os.sep)
                    year = next((part for part in reversed(path_parts) if part.isdigit() and len(part) == 4), None)
                    if not year:
                        continue

                    try:
                        _, date_dt = self._parse_file_date(year, filename)
                    except Exception:
                        continue

                    file_type = filename_lower.split(".")[-1]
                    file_path = os.path.join(root, filename)

                    current = latest_by_type.get(file_type)
                    if current is None or date_dt > current["date"]:
                        latest_by_type[file_type] = {
                            "file_path": file_path,
                            "date": date_dt,
                            "date_str": date_dt.strftime("%Y-%m-%d"),
                        }

            for file_type in ["txf", "txr", "tx2"]:
                file_info = latest_by_type.get(file_type)
                if not file_info:
                    continue
                last_demand_document = db.query(LastDemandDocument).filter_by(
                    document_type=file_type.lower()
                ).first()
                if last_demand_document:
                    last_demand_document.document_route = file_info["file_path"]
                    last_demand_document.document_date = file_info["date"]
                else:
                    new_demand_document = LastDemandDocument(
                        document_type=file_type.lower(),
                        document_route=file_info["file_path"],
                        document_date=file_info["date"]
                    )
                    db.add(new_demand_document)

            db.commit()

            return {
                "message": "Documents verified from filesystem",
                "last_txf_file": latest_by_type.get("txf", {}).get("file_path"),
                "last_txr_file": latest_by_type.get("txr", {}).get("file_path"),
                "last_tx2_file": latest_by_type.get("tx2", {}).get("file_path"),
            }
        except Exception as e:
            print(f"Error updating LastDemandDocument from filesystem: {str(e)}")
            db.rollback()
            raise e

    def update_last_demand_documents_from_ftp(self, db: Session):
        """
        Actualiza la tabla LastDemandDocument basandose en archivos ADEM disponibles en FTP.
        """
        try:
            from .data import GMRAdemService

            ftp_service = GMRAdemService()
            ftp_contents = ftp_service.list_ftp_adem_directory()
            if not ftp_contents.get("success", False):
                raise Exception(f"Error listing FTP ADEM directory: {ftp_contents.get('error', 'Unknown error')}")

            directories = ftp_contents.get("directory_contents", {}).get("directories", [])
            latest_by_type = {}

            for directory in directories:
                month_dir = directory.get("name")
                if not month_dir or len(month_dir) != 7 or "-" not in month_dir:
                    continue

                year_part, month_part = month_dir.split("-", 1)
                if not (year_part.isdigit() and month_part.isdigit()):
                    continue

                try:
                    files = ftp_service.list_adem_files_from_month(month_dir)
                except Exception:
                    continue

                for filename in files:
                    filename_lower = filename.lower()
                    if not filename_lower.startswith("adem"):
                        continue

                    try:
                        _, date_dt = self._parse_file_date(year_part, filename)
                    except Exception:
                        continue

                    file_type = filename_lower.split(".")[-1]
                    file_path = f"{month_dir}/ADEM/{filename}"

                    current = latest_by_type.get(file_type)
                    if current is None or date_dt > current["date"]:
                        latest_by_type[file_type] = {
                            "file_path": file_path,
                            "date": date_dt,
                            "date_str": date_dt.strftime("%Y-%m-%d"),
                        }

            for file_type in ["txf", "txr", "tx2"]:
                file_info = latest_by_type.get(file_type)
                if not file_info:
                    continue
                last_demand_document = db.query(LastDemandDocument).filter_by(
                    document_type=file_type.lower()
                ).first()
                if last_demand_document:
                    last_demand_document.document_route = file_info["file_path"]
                    last_demand_document.document_date = file_info["date"]
                else:
                    new_demand_document = LastDemandDocument(
                        document_type=file_type.lower(),
                        document_route=file_info["file_path"],
                        document_date=file_info["date"]
                    )
                    db.add(new_demand_document)

            db.commit()

            return {
                "message": "Documents verified from ftp",
                "last_txf_file": latest_by_type.get("txf", {}).get("file_path"),
                "last_txr_file": latest_by_type.get("txr", {}).get("file_path"),
                "last_tx2_file": latest_by_type.get("tx2", {}).get("file_path"),
            }
        except Exception as e:
            print(f"Error updating LastDemandDocument from ftp: {str(e)}")
            db.rollback()
            raise e

    def verify_documents(self, db: Session):
            try:
                # Solo leer desde la tabla GMR_last_demand_document
                return self.get_latest_files_by_type(db)

            except SQLAlchemyError as e:
                db.rollback()
                raise e
            except Exception as e:
                raise e

    def get_latest_files_by_type(self, db: Session):
        """
        Obtiene los registros más recientes de cada tipo de archivo desde LastDemandDocument
        """
        try:
            latest_files = {}
            
            for file_type in ['txf', 'txr', 'tx2']:
                last_doc = db.query(LastDemandDocument).filter_by(
                    document_type=file_type.lower()
                ).first()
                if last_doc and last_doc.document_date:
                    latest_files[file_type] = {
                        'file_path': last_doc.document_route,
                        'date': last_doc.document_date,
                        'date_str': last_doc.document_date.strftime('%Y-%m-%d')
                    }
                    print(f"Último archivo {file_type.upper()}: {last_doc.document_route} (fecha: {last_doc.document_date.strftime('%Y-%m-%d')})")
                else:
                    print(f"No se encontraron registros de tipo {file_type.upper()}")
            
            return latest_files
            
        except Exception as e:
            print(f"Error getting latest files by type: {str(e)}")
            return {}

    def get_demand_from_db_by_date_range(self, db: Session, start_date: str, end_date: str):
        """
        Método auxiliar para obtener demandas en un rango de fechas específico
        """
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            
            demands = db.query(Demand).filter(
                Demand.fecha >= start_dt,
                Demand.fecha <= end_dt
            ).order_by(Demand.fecha).all()
            
            daily_demands = [round(demand.total, 2) for demand in demands]
            dates = [demand.fecha.strftime('%Y-%m-%d') for demand in demands]
            
            return {
                "daily_demands": daily_demands,
                "dates": dates
            }
            
        except Exception as e:
            raise Exception(f"Error obtaining demand data from DB: {str(e)}")

    def load_model(self, folder='api/models/xgboost'):
        """Carga el modelo y sus variables"""
        import sys
        import pickle
        from api.services.functions.xgb_temp import DemandPredictor
        
        # Solución para el problema de carga con joblib
        # Añadir el módulo correcto al sys.modules con el nombre que busca joblib
        sys.modules['xgBoost'] = sys.modules['api.services.functions.xgb_temp']
        
        predictor = DemandPredictor()
        
        try:
            # Cargar el modelo usando pickle directamente si joblib falla
            try:
                predictor.model = joblib.load(f'{folder}/model.joblib')
            except Exception as e:
                print(f"Fallback a pickle para cargar el modelo: {str(e)}")
                with open(f'{folder}/model.joblib', 'rb') as f:
                    predictor.model = pickle.load(f)
            
            predictor.label_encoder = joblib.load(f'{folder}/label_encoder.joblib')
            predictor.scale_params = joblib.load(f'{folder}/scale_params.joblib')
            
        except Exception as e:
            print(f"Error cargando el modelo: {str(e)}")
            raise e
        
        return predictor

    def load_model_varias(self, folder='api/models/xgboost_varias'):
        """Carga el modelo y sus variables"""
        from api.services.functions.xgb_varias import DemandVariasPredictor
        import sys
        import pickle

        # Solución para el problema de carga con joblib
        # Añadir el módulo correcto al sys.modules con el nombre que busca joblib
        sys.modules['xgBoost_varias'] = sys.modules['api.services.functions.xgb_varias']
        
        predictor = DemandVariasPredictor()
        
        try:
            # Cargar el modelo usando pickle directamente si joblib falla
            try:
                predictor.model = joblib.load(f'{folder}/model.joblib')
            except Exception as e:
                print(f"Fallback a pickle para cargar el modelo: {str(e)}")
                with open(f'{folder}/model.joblib', 'rb') as f:
                    predictor.model = pickle.load(f)
                
            predictor.label_encoder = joblib.load(f'{folder}/label_encoder.joblib')
            predictor.scale_params = joblib.load(f'{folder}/scale_params.joblib')
            
        except Exception as e:
            print(f"Error cargando el modelo: {str(e)}")
            raise e
        
        return predictor

    def load_model_only(self, folder='api/models/xgboost_only'):
        """Carga el modelo y sus variables"""
        import sys
        import pickle
        from api.services.functions.xgb_only import DemandOnlyPredictor

        # Solución para el problema de carga con joblib
        # Añadir el módulo correcto al sys.modules con el nombre que busca joblib
        sys.modules['xgBoost_only'] = sys.modules['api.services.functions.xgb_only']
        
        predictor = DemandOnlyPredictor()
        
        try:
            # Cargar el modelo usando pickle directamente si joblib falla
            try:
                predictor.model = joblib.load(f'{folder}/model.joblib')
            except Exception as e:
                print(f"Fallback a pickle para cargar el modelo: {str(e)}")
                with open(f'{folder}/model.joblib', 'rb') as f:
                    predictor.model = pickle.load(f)
                    
            # Cargar los parámetros de escala
            predictor.scale_params = joblib.load(f'{folder}/scale_params.joblib')
            
        except Exception as e:
            print(f"Error cargando el modelo: {str(e)}")
            raise e
        
        return predictor

    def load_model_multimodelo(self, folder='api/models/xgboost_multimodelo'):
        """Carga el modelo multimodelo con 3 submodelos especializados"""
        import sys
        import pickle
        from api.services.functions.xgBoost_multimodelo import DemandPredictorMultiModelo

        # Solución para el problema de carga con joblib
        sys.modules['xgBoost_multimodelo'] = sys.modules['api.services.functions.xgBoost_multimodelo']

        predictor = DemandPredictorMultiModelo()

        try:
            predictor.load_models(base_path=folder)
        except Exception as e:
            print(f"Error cargando modelo multimodelo: {str(e)}")
            raise e

        return predictor

    def load_model_varias_multimodelo(self, folder='api/models/xgboost_varias_multimodelo'):
        """Carga el modelo multimodelo de 3 temperaturas"""
        import sys
        import pickle
        from api.services.functions.xgBoost_varias_multimodelo import DemandVariasMultiModelo

        # Solución para el problema de carga con joblib
        sys.modules['xgBoost_varias_multimodelo'] = sys.modules['api.services.functions.xgBoost_varias_multimodelo']

        predictor = DemandVariasMultiModelo()

        try:
            predictor.load_models(base_path=folder)
        except Exception as e:
            print(f"Error cargando modelo varias multimodelo: {str(e)}")
            raise e

        return predictor

    def get_type_day(self, date_list, year):
        # Get holidays for the specified year
        holidays = get_colombia_holidays_by_year(year)
        # Convert to list and extract dates as strings
        holidays = list(holidays)
        holidays_dates = [holiday.date.strftime('%Y-%m-%d') for holiday in holidays]
        
        day_names = {
            0: "Lunes",
            1: "Martes", 
            2: "Miércoles",
            3: "Jueves",
            4: "Viernes",
            5: "Sábado",
            6: "Domingo"
        }
        
        day_types = []
        for date_str in date_list:
            date_obj = datetime.strptime(date_str, '%Y-%m-%d')
            weekday = date_obj.weekday()
            day_name = day_names[weekday]
            
            if date_str in holidays_dates:
                day_types.append(f"{day_name} festivo")
            else:
                day_types.append(day_name)
            
        return day_types

    def _get_preferred_demands_for_range(self, db: Session, start_date: datetime, end_date: datetime):
        """
        Devuelve demandas por fecha usando prioridad por dia (txf > txr > tx2).
        """
        priority_case = case(
            (Demand.file_type == "txf", 0),
            (Demand.file_type == "txr", 1),
            (Demand.file_type == "tx2", 2),
            else_=3,
        )
        rows = (
            db.query(Demand)
            .filter(Demand.fecha >= start_date, Demand.fecha <= end_date)
            .order_by(Demand.fecha.asc(), priority_case.asc())
            .all()
        )
        demands_by_date = {}
        for row in rows:
            date_key = row.fecha.strftime("%Y-%m-%d")
            if date_key not in demands_by_date:
                demands_by_date[date_key] = row
        return [demands_by_date[k] for k in sorted(demands_by_date.keys())]


    def get_demand(self, year: int, month: int, previous_days: int, db: Session):   
        """
        Obtiene datos de demanda desde la base de datos.
        previous_days: número entre 0 y 15 que indica cuántos días previos incluir
        """
        try:
            daily_demands = []
            dates = []

            # Validar que previous_days esté en el rango correcto
            if previous_days < 0 or previous_days > 15:
                raise Exception("previous_days debe estar entre 0 y 15")

            # Calcular fechas de inicio y fin del mes actual
            current_month_start = datetime(year, month, 1)
            current_month_end = datetime(year, month, calendar.monthrange(year, month)[1])
            
            # Si previous_days > 0, incluir los días previos especificados del mes anterior
            if previous_days > 0:
                if month == 1:
                    previous_year = year - 1
                    previous_month = 12
                else:
                    previous_year = year
                    previous_month = month - 1
                
                # Obtener el último día del mes anterior
                previous_month_last_day = calendar.monthrange(previous_year, previous_month)[1]
                
                # Calcular fecha de inicio basada en previous_days
                if previous_month_last_day >= previous_days:
                    start_day = previous_month_last_day - previous_days + 1
                else:
                    start_day = 1
                
                previous_start = datetime(previous_year, previous_month, start_day)
                previous_end = datetime(previous_year, previous_month, previous_month_last_day)
                
                # Consultar demandas del mes anterior usando prioridad por dia (txf > txr > tx2)
                previous_demands = self._get_preferred_demands_for_range(
                    db, previous_start, previous_end
                )
                
                # Filtrar solo los últimos previous_days días si hay más registros
                if len(previous_demands) > previous_days:
                    previous_demands = previous_demands[-previous_days:]
                
                for demand in previous_demands:
                    daily_demands.append(round(demand.total, 2))
                    dates.append(demand.fecha.strftime('%Y-%m-%d'))

            # Consultar demandas del mes actual usando prioridad por dia (txf > txr > tx2)
            current_demands = self._get_preferred_demands_for_range(
                db, current_month_start, current_month_end
            )
            
            for demand in current_demands:
                daily_demands.append(round(demand.total, 2))
                dates.append(demand.fecha.strftime('%Y-%m-%d'))

            # Obtener tipos de día para todas las fechas
            day_types = self.get_type_day(dates, year)
            
            climate = ClimateService()
            clima_ciudades = {}
            
            if dates:  # Solo si hay fechas
                # Definir las 4 ciudades principales
                ciudades_principales = ["Cesar", "CordobaSucre", "Bolivar"]
                
                # Crear search object para el clima
                search_temp = SearchAllClimateTypeSchema(
                    fecha_inicio=dates[0],
                    fecha_fin=dates[-1],
                    id=0  # id=0 para temperaturas
                )
                
                # Obtener datos para cada ciudad
                for ciudad in ciudades_principales:
                    try:
                        city_data = climate.get_climate_daily_all_types_per_city(db=db, search=search_temp, city=ciudad)
                        
                        # Encontrar el índice de inicio en los datos de la ciudad
                        city_start_index = 0
                        if dates and city_data["fechas"]:
                            for i, city_date in enumerate(city_data["fechas"]):
                                if city_date == dates[0]:
                                    city_start_index = i
                                    break
                        
                        # Extraer datos correspondientes a nuestras fechas
                        data_count = len(dates)
                        
                        # Extraer temperaturas
                        temp_max = city_data["temperaturas"]["maximas"][city_start_index:city_start_index + data_count]
                        temp_min = city_data["temperaturas"]["minimas"][city_start_index:city_start_index + data_count]
                        temp_avg = city_data["temperaturas"]["promedios"][city_start_index:city_start_index + data_count]
                        
                        # Extraer humedades
                        hum_max = city_data["humedades"]["maximas"][city_start_index:city_start_index + data_count]
                        hum_min = city_data["humedades"]["minimas"][city_start_index:city_start_index + data_count]
                        hum_avg = city_data["humedades"]["promedios"][city_start_index:city_start_index + data_count]
                        
                        # Extraer velocidades
                        vel_max = city_data["velocidades"]["maximas"][city_start_index:city_start_index + data_count]
                        vel_min = city_data["velocidades"]["minimas"][city_start_index:city_start_index + data_count]
                        vel_avg = city_data["velocidades"]["promedios"][city_start_index:city_start_index + data_count]
                        
                        # Estructura simple para el frontend
                        clima_ciudades[ciudad.lower().replace("cordobasucre", "cordoba_sucre")] = {
                            "temp_max": [float(x) if x is not None else None for x in temp_max],
                            "temp_min": [float(x) if x is not None else None for x in temp_min],
                            "temp_avg": [float(x) if x is not None else None for x in temp_avg],
                            "hum_max": [float(x) if x is not None else None for x in hum_max],
                            "hum_min": [float(x) if x is not None else None for x in hum_min],
                            "hum_avg": [float(x) if x is not None else None for x in hum_avg],
                            "vel_max": [float(x) if x is not None else None for x in vel_max],
                            "vel_min": [float(x) if x is not None else None for x in vel_min],
                            "vel_avg": [float(x) if x is not None else None for x in vel_avg]
                        }
                        
                    except Exception as e:
                        print(f"Error obteniendo datos de {ciudad}: {str(e)}")
                        clima_ciudades[ciudad.lower().replace("cordobasucre", "cordoba_sucre")] = {
                            "temp_max": [None] * len(dates),
                            "temp_min": [None] * len(dates),
                            "temp_avg": [None] * len(dates),
                            "hum_max": [None] * len(dates),
                            "hum_min": [None] * len(dates),
                            "hum_avg": [None] * len(dates),
                            "vel_max": [None] * len(dates),
                            "vel_min": [None] * len(dates),
                            "vel_avg": [None] * len(dates)
                        }
                
                if "Bolivar" in clima_ciudades:
                    temp_values = clima_ciudades["bolivar"]["temp_avg"]
                    temp_max_values = clima_ciudades["bolivar"]["temp_max"]
                    temp_min_values = clima_ciudades["bolivar"]["temp_min"]
                else:
                    temp_values = [None] * len(dates)
                    temp_max_values = [None] * len(dates)
                    temp_min_values = [None] * len(dates)
            else:
                # Si no hay fechas, devolver estructuras vacías
                temp_values = []
                temp_max_values = []
                temp_min_values = []
                clima_ciudades = {}

            return {
                "daily_demands": daily_demands,
                "dates": dates,
                "day_types": day_types,
                "temp": [float(x) if x is not None else None for x in temp_values],
                "max_temp": [float(x) if x is not None else None for x in temp_max_values],
                "min_temp": [float(x) if x is not None else None for x in temp_min_values],
                "clima": clima_ciudades
            }
        except Exception as e:
            raise Exception(f"Exception on get_demand: {str(e)}")

    def predict(self, search: MPMPredictSchema, db: Session):
        try:
            date = search.last_date
            year = int(date[:4])
            month = int(date[5:7])
            
            

            # Get the last day of the month using calendar module
            _, last_day = calendar.monthrange(year, month)
            month_days = calendar.monthrange(year, month)[1]
            last_month_date = f"{year}-{month:02d}-{last_day:02d}"
            
            # Obtener datos históricos desde la BD (incluye días previos para cálculos)
            historic_data = self.get_demand(year, month, search.previous_days, db)
            historic_demands = historic_data["daily_demands"]
            all_dates = historic_data["dates"]
            
            if len(historic_demands) < 1:
                raise Exception("No hay datos históricos para realizar la predicción")
            
            # Separar datos del mes actual de los días previos
            current_month_start = datetime(year, month, 1)
            
            current_month_demands = []
            current_month_dates = []
            
            for i, date_str in enumerate(all_dates):
                date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                if date_obj >= current_month_start:
                    current_month_demands.append(historic_demands[i])
                    current_month_dates.append(date_str)
            
            # Crear array de fechas de predicción
            if current_month_dates:
                last_historic_date = datetime.strptime(current_month_dates[-1], "%Y-%m-%d")
            else:
                last_historic_date = current_month_start - pd.Timedelta(days=1)
                
            prediction_dates_array = []
            current_date = last_historic_date + pd.Timedelta(days=1)
            end_date = datetime.strptime(last_month_date, "%Y-%m-%d")
            
            while current_date <= end_date:
                prediction_dates_array.append(current_date.strftime("%Y-%m-%d"))
                current_date += pd.Timedelta(days=1)
            
            # Obtener datos de clima ANTES de verificar si hay predicciones
            results = {}
            climate = ClimateService()
            search_temp = SearchAllClimateTypeSchema(
                fecha_inicio=all_dates[0],  # Usar fecha desde días previos para obtener temperatura
                fecha_fin=last_month_date,
                id=0)
            temp_data = climate.get_climate_day_all_types(db=db, search=search_temp)
            print("fechas temperatura", temp_data["lista_fechas"])
            temp_list = temp_data["lista_avg"]
            temp_list_max = temp_data["lista_max"]
            temp_list_min = temp_data["lista_min"]

            if not prediction_dates_array:
                # Si no hay predicciones, aún debemos devolver day_types y temperaturas para los históricos
                historic_day_types = self.get_type_day(current_month_dates, year)
                
                # Encontrar temperaturas para los días históricos del mes
                temp_month_start_index = 0
                for i, temp_date in enumerate(temp_data["lista_fechas"]):
                    if temp_date == current_month_dates[0]:
                        temp_month_start_index = i
                        break
                
                historic_days_count = len(current_month_dates)
                historic_temp = temp_list[temp_month_start_index:temp_month_start_index + historic_days_count]
                historic_temp_max = temp_list_max[temp_month_start_index:temp_month_start_index + historic_days_count]
                historic_temp_min = temp_list_min[temp_month_start_index:temp_month_start_index + historic_days_count]

                clima_ciudades = {}
                if current_month_dates:
                    ciudades_principales = ["Cesar", "CordobaSucre", "Bolivar"]
                    search_temp_cities = SearchAllClimateTypeSchema(
                        fecha_inicio=current_month_dates[0],
                        fecha_fin=current_month_dates[-1],
                        id=0
                    )

                    for ciudad in ciudades_principales:
                        try:
                            city_data = climate.get_climate_daily_all_types_per_city(
                                db=db,
                                search=search_temp_cities,
                                city=ciudad
                            )

                            city_start_index = 0
                            if current_month_dates and city_data["fechas"]:
                                for i, city_date in enumerate(city_data["fechas"]):
                                    if city_date == current_month_dates[0]:
                                        city_start_index = i
                                        break

                            data_count = len(current_month_dates)
                            temp_max = city_data["temperaturas"]["maximas"][city_start_index:city_start_index + data_count]
                            temp_min = city_data["temperaturas"]["minimas"][city_start_index:city_start_index + data_count]
                            temp_avg = city_data["temperaturas"]["promedios"][city_start_index:city_start_index + data_count]

                            hum_max = city_data["humedades"]["maximas"][city_start_index:city_start_index + data_count]
                            hum_min = city_data["humedades"]["minimas"][city_start_index:city_start_index + data_count]
                            hum_avg = city_data["humedades"]["promedios"][city_start_index:city_start_index + data_count]

                            vel_max = city_data["velocidades"]["maximas"][city_start_index:city_start_index + data_count]
                            vel_min = city_data["velocidades"]["minimas"][city_start_index:city_start_index + data_count]
                            vel_avg = city_data["velocidades"]["promedios"][city_start_index:city_start_index + data_count]

                            clima_ciudades[ciudad.lower().replace("cordobasucre", "cordoba_sucre")] = {
                                "temp_max": [float(x) if x is not None else None for x in temp_max],
                                "temp_min": [float(x) if x is not None else None for x in temp_min],
                                "temp_avg": [float(x) if x is not None else None for x in temp_avg],
                                "hum_max": [float(x) if x is not None else None for x in hum_max],
                                "hum_min": [float(x) if x is not None else None for x in hum_min],
                                "hum_avg": [float(x) if x is not None else None for x in hum_avg],
                                "vel_max": [float(x) if x is not None else None for x in vel_max],
                                "vel_min": [float(x) if x is not None else None for x in vel_min],
                                "vel_avg": [float(x) if x is not None else None for x in vel_avg]
                            }
                        except Exception as e:
                            print(f"Error obteniendo datos de {ciudad} en predict: {str(e)}")
                            clima_ciudades[ciudad.lower().replace("cordobasucre", "cordoba_sucre")] = {
                                "temp_max": [None] * len(current_month_dates),
                                "temp_min": [None] * len(current_month_dates),
                                "temp_avg": [None] * len(current_month_dates),
                                "hum_max": [None] * len(current_month_dates),
                                "hum_min": [None] * len(current_month_dates),
                                "hum_avg": [None] * len(current_month_dates),
                                "vel_max": [None] * len(current_month_dates),
                                "vel_min": [None] * len(current_month_dates),
                                "vel_avg": [None] * len(current_month_dates)
                            }
                
                return {
                    "message": "No hay días para predecir. Todos los datos del mes ya están disponibles.",
                    "historic_demands": [float(x) for x in current_month_demands],
                    "prediction": [],
                    "dates": current_month_dates,
                    "day_types": historic_day_types,  # Tipos de día para históricos
                    "temp": [float(x) if x is not None else None for x in historic_temp],
                    "max_temp": [float(x) if x is not None else None for x in historic_temp_max],
                    "min_temp": [float(x) if x is not None else None for x in historic_temp_min],
                    "total_predictions": float(sum(current_month_demands)),
                    "clima": clima_ciudades
                }
            
            # Crear array completo de fechas (históricos + predicciones)
            all_dates_complete = current_month_dates + prediction_dates_array
            
            # Generate day types for ALL dates (historic + predicted) del mes completo
            day_types = self.get_type_day(all_dates_complete, year)
            print("day_types", day_types)
            print("historic_demands", historic_demands)
            print("current_month_demands", current_month_demands)
            print("dates", all_dates)
            print("current_month_dates", current_month_dates)
            print("all_dates_complete", all_dates_complete)
            
            days_to_predict = len(prediction_dates_array)
            print("days_to_predict", days_to_predict)
            
            # Demand only prediction
            if search.id == 1:
                predictions = []
                predictor = self.load_model_only()
                
                # Usar TODOS los datos históricos (incluyendo días previos) para el modelo
                df = pd.DataFrame({
                    "Fecha": pd.to_datetime(all_dates, format='%Y-%m-%d'),
                    "P": historic_demands
                })
                
                df.to_csv("mpm_prueba.csv")

                for i in range(1, days_to_predict + 1):
                    prediction, _ = predictor.predict(df)
                    
                    last_date = df["Fecha"].iloc[-1]
                    correct_prediction_date = last_date + pd.Timedelta(days=1)
                    
                    # Convertir predicción a float de Python
                    predictions.append(float(prediction))
                    new_row = pd.DataFrame({"Fecha": [correct_prediction_date], "P": [float(prediction)]})
                    df = pd.concat([df, new_row], ignore_index=True)

            # Temperature and demand prediction
            elif search.id == 2:
                # Determinar tipo de día basado en search.type
                if search.type == 0:
                    day_type = "caluroso"
                elif search.type == 1:
                    day_type = "templado"
                elif search.type == 2:
                    day_type = "fresco"
                else:
                    day_type = "templado"  # valor por defecto
                    
                predictions = []
                predictor = self.load_model_multimodelo()

                # Find temperature index offset (diferencia entre fechas de temperatura y fechas de demanda)
                temp_dates = temp_data["lista_fechas"]
                temp_start_index = 0
                
                for i, temp_date in enumerate(temp_dates):
                    if temp_date == all_dates[0]:
                        temp_start_index = i
                        break
                
                # Usar TODOS los datos históricos (incluyendo días previos) para el modelo
                df = pd.DataFrame({
                    "Fecha": pd.to_datetime(all_dates, format='%Y-%m-%d'),
                    "P": historic_demands
                })
                
                # Agregar temperatura correspondiente
                for i in range(len(historic_demands)):
                    temp_index = temp_start_index + i
                    if temp_index < len(temp_list):
                        # Usar siempre temperatura promedio para el método 2 (una temperatura)
                        df.loc[i, "t"] = temp_list[temp_index]
                
                # Para las predicciones futuras, usar temperatura promedio
                temp_list_to_predict = temp_list[len(historic_demands):]

                for i in range(1, days_to_predict + 1):
                    # Pasar day_type como argumento requerido
                    prediction, _ = predictor.predict(df, day_type)
                    
                    last_date = df["Fecha"].iloc[-1]
                    correct_prediction_date = last_date + pd.Timedelta(days=1)
                    
                    # Convertir predicción a float de Python
                    predictions.append(float(prediction))
                    
                    new_row = pd.DataFrame({
                        "Fecha": [correct_prediction_date], 
                        "P": [float(prediction)], 
                        "t": [temp_list_to_predict[i-1]]
                    })
                    df = pd.concat([df, new_row], ignore_index=True)

            # Three temperatures prediction

            elif search.id == 3:
                # Determinar tipo de día basado en search.type
                if search.type == 0:
                    day_type = "caluroso"
                elif search.type == 1:
                    day_type = "templado"
                elif search.type == 2:
                    day_type = "fresco"
                else:
                    day_type = "templado"  # valor por defecto
                    
                predictions = []
                # CORRECCIÓN: Usar load_model_varias_multimodelo para variación climática
                predictor = self.load_model_varias_multimodelo()

                # ===== NUEVA SECCIÓN: Obtener datos por ciudad =====
                climate = ClimateService()
                
                # Obtener temperaturas específicas de las 3 ciudades
                search_temp_cities = SearchAllClimateTypeSchema(
                    fecha_inicio=all_dates[0],
                    fecha_fin=last_month_date,
                    id=0  # id=0 para temperaturas
                )
                
                # Bolívar para t_CTG
                bolivar_data = climate.get_climate_daily_all_types_per_city(db=db, search=search_temp_cities, city="Bolivar")
                temp_bolivar = bolivar_data["temperaturas"]["promedios"]
                
                # Barranquilla para t_MON
                barranquilla_data = climate.get_climate_daily_all_types_per_city(db=db, search=search_temp_cities, city="CordobaSucre")
                temp_barranquilla = barranquilla_data["temperaturas"]["promedios"]
                
                # Cesar para t_VDP
                cesar_data = climate.get_climate_daily_all_types_per_city(db=db, search=search_temp_cities, city="Cesar")
                temp_cesar = cesar_data["temperaturas"]["promedios"]
                
                # Encontrar índice de inicio para las temperaturas de las ciudades
                temp_start_index = 0
                for i, temp_date in enumerate(bolivar_data["fechas"]):
                    if temp_date == all_dates[0]:
                        temp_start_index = i
                        break
                
                # Usar TODOS los datos históricos (incluyendo días previos) para el modelo
                df = pd.DataFrame({
                    "Fecha": pd.to_datetime(all_dates, format='%Y-%m-%d'),
                    "P": historic_demands
                })
                
                # Agregar las tres temperaturas usando datos específicos de cada ciudad
                for i in range(len(historic_demands)):
                    temp_index = temp_start_index + i
                    if temp_index < len(temp_bolivar):
                        df.loc[i, "t_CTG"] = temp_bolivar[temp_index]    # Bolívar
                        df.loc[i, "t_MON"] = temp_barranquilla[temp_index]  # Barranquilla
                        df.loc[i, "t_VDP"] = temp_cesar[temp_index]      # Cesar
                
                # Preparar temperaturas para predicciones futuras
                temp_list_to_predict_ctg = temp_bolivar[len(historic_demands):]      # Bolívar
                temp_list_to_predict_mon = temp_barranquilla[len(historic_demands):] # Barranquilla
                temp_list_to_predict_vdp = temp_cesar[len(historic_demands):]        # Cesar

                for i in range(1, days_to_predict + 1):
                    # Pasar day_type como argumento requerido
                    prediction, _ = predictor.predict(df, day_type)
                    
                    last_date = df["Fecha"].iloc[-1]
                    correct_prediction_date = last_date + pd.Timedelta(days=1)
                    
                    # Convertir predicción a float de Python
                    predictions.append(float(prediction))
                    new_row = pd.DataFrame({
                        "Fecha": [correct_prediction_date], 
                        "P": [float(prediction)], 
                        "t_CTG": [temp_list_to_predict_ctg[i-1]],  # Bolívar
                        "t_MON": [temp_list_to_predict_mon[i-1]],  # Barranquilla
                        "t_VDP": [temp_list_to_predict_vdp[i-1]]   # Cesar
                    })
                    df = pd.concat([df, new_row], ignore_index=True)

            # Average demand prediction
            elif search.id == 4:
                predictions = []
                ordinary_days_demands = []
                saturdays_demands = []
                sundays_and_holidays_demands = []
                
                # Obtener los tipos de día para las fechas históricas (TODOS los datos)
                historic_day_types = self.get_type_day(all_dates, year)
                
                # Clasificar los datos históricos (usar TODOS para calcular promedios)
                for i in range(len(historic_demands)):
                    if historic_day_types[i] == "Sábado":
                        saturdays_demands.append(historic_demands[i])
                    elif historic_day_types[i] == "Domingo" or "festivo" in historic_day_types[i]:
                        sundays_and_holidays_demands.append(historic_demands[i])
                    else:
                        ordinary_days_demands.append(historic_demands[i])
                
                # Calcular promedios y convertir a float de Python
                ordinary_demand = float(np.mean(ordinary_days_demands)) if ordinary_days_demands else 0.0
                saturday_demand = float(np.mean(saturdays_demands)) if saturdays_demands else 0.0
                sunday_demand = float(np.mean(sundays_and_holidays_demands)) if sundays_and_holidays_demands else 0.0
                
                print("ordinary_demand", ordinary_demand)
                print("saturday_demand", saturday_demand)
                print("sunday_demand", sunday_demand)
                

                df = pd.DataFrame({
                    "Fecha": pd.to_datetime(all_dates, format='%Y-%m-%d'),
                    "P": historic_demands
                })

                # Obtener los tipos de día para las fechas de predicción
                future_day_types = self.get_type_day(prediction_dates_array, year)
                
                # Para cada día futuro a predecir
                for i in range(days_to_predict):
                    last_date = df["Fecha"].iloc[-1]
                    correct_prediction_date = last_date + pd.Timedelta(days=1)
                    
                    # Obtener el tipo de día para la fecha de predicción
                    current_day_type = future_day_types[i]
                    
                    # Determinar la predicción basada en el tipo de día y convertir a float
                    if "Sábado" in current_day_type:
                        prediction = float(saturday_demand)
                    elif "Domingo" in current_day_type or "festivo" in current_day_type:
                        prediction = float(sunday_demand)
                    else:
                        prediction = float(ordinary_demand)
                    
                    predictions.append(prediction)
                    new_row = pd.DataFrame({"Fecha": [correct_prediction_date], "P": [prediction]})
                    df = pd.concat([df, new_row], ignore_index=True)

            print(df)
            
            if current_month_dates:
                month_reference_date = current_month_dates[0]
            else:
                month_reference_date = prediction_dates_array[0]

            temp_month_start_index = 0
            for i, temp_date in enumerate(temp_data["lista_fechas"]):
                if temp_date == month_reference_date:  # Primera fecha del mes actual
                    temp_month_start_index = i
                    break
            
            # Extraer temperaturas para todo el mes (históricos del mes + predicciones)
            total_days_in_month = len(all_dates_complete)
            temp_complete = temp_list[temp_month_start_index:temp_month_start_index + total_days_in_month]
            temp_max_complete = temp_list_max[temp_month_start_index:temp_month_start_index + total_days_in_month]
            temp_min_complete = temp_list_min[temp_month_start_index:temp_month_start_index + total_days_in_month]
            
            clima_ciudades = {}
            
            # Definir las 4 ciudades principales
            ciudades_principales = ["Cesar", "CordobaSucre", "Bolivar"]
            
            # Crear search object para el clima por ciudad
            search_temp_cities = SearchAllClimateTypeSchema(
                fecha_inicio=all_dates_complete[0],
                fecha_fin=all_dates_complete[-1],
                id=0  # id=0 para temperaturas
            )
            
            # Obtener datos para cada ciudad
            for ciudad in ciudades_principales:
                try:
                    city_data = climate.get_climate_daily_all_types_per_city(db=db, search=search_temp_cities, city=ciudad)
                    
                    # Encontrar el índice de inicio en los datos de la ciudad
                    city_start_index = 0
                    if all_dates_complete and city_data["fechas"]:
                        for i, city_date in enumerate(city_data["fechas"]):
                            if city_date == all_dates_complete[0]:
                                city_start_index = i
                                break
                    
                    # Extraer datos correspondientes al período completo del mes
                    data_count = len(all_dates_complete)
                    
                    # Extraer temperaturas
                    temp_max = city_data["temperaturas"]["maximas"][city_start_index:city_start_index + data_count]
                    temp_min = city_data["temperaturas"]["minimas"][city_start_index:city_start_index + data_count]
                    temp_avg = city_data["temperaturas"]["promedios"][city_start_index:city_start_index + data_count]
                    
                    # Extraer humedades
                    hum_max = city_data["humedades"]["maximas"][city_start_index:city_start_index + data_count]
                    hum_min = city_data["humedades"]["minimas"][city_start_index:city_start_index + data_count]
                    hum_avg = city_data["humedades"]["promedios"][city_start_index:city_start_index + data_count]
                    
                    # Extraer velocidades
                    vel_max = city_data["velocidades"]["maximas"][city_start_index:city_start_index + data_count]
                    vel_min = city_data["velocidades"]["minimas"][city_start_index:city_start_index + data_count]
                    vel_avg = city_data["velocidades"]["promedios"][city_start_index:city_start_index + data_count]
                    
                    # Estructura simple para el frontend
                    clima_ciudades[ciudad.lower().replace("cordobasucre", "cordoba_sucre")] = {
                        "temp_max": [float(x) if x is not None else None for x in temp_max],
                        "temp_min": [float(x) if x is not None else None for x in temp_min],
                        "temp_avg": [float(x) if x is not None else None for x in temp_avg],
                        "hum_max": [float(x) if x is not None else None for x in hum_max],
                        "hum_min": [float(x) if x is not None else None for x in hum_min],
                        "hum_avg": [float(x) if x is not None else None for x in hum_avg],
                        "vel_max": [float(x) if x is not None else None for x in vel_max],
                        "vel_min": [float(x) if x is not None else None for x in vel_min],
                        "vel_avg": [float(x) if x is not None else None for x in vel_avg]
                    }
                    
                except Exception as e:
                    print(f"Error obteniendo datos de {ciudad} en predict: {str(e)}")
                    clima_ciudades[ciudad.lower().replace("cordobasucre", "cordoba_sucre")] = {
                        "temp_max": [None] * len(all_dates_complete),
                        "temp_min": [None] * len(all_dates_complete),
                        "temp_avg": [None] * len(all_dates_complete),
                        "hum_max": [None] * len(all_dates_complete),
                        "hum_min": [None] * len(all_dates_complete),
                        "hum_avg": [None] * len(all_dates_complete),
                        "vel_max": [None] * len(all_dates_complete),
                        "vel_min": [None] * len(all_dates_complete),
                        "vel_avg": [None] * len(all_dates_complete)
                    }
            
            results["historic_demands"] = [float(x) for x in current_month_demands] 
            results["prediction"] = predictions  
            results["dates"] = all_dates_complete 
            results["day_types"] = day_types  
            
            # Mantener campos originales para compatibilidad (ahora basados en Barranquilla)
            if "bolivar" in clima_ciudades:
                results["temp"] = clima_ciudades["bolivar"]["temp_avg"]
                results["max_temp"] = clima_ciudades["bolivar"]["temp_max"]
                results["min_temp"] = clima_ciudades["bolivar"]["temp_min"]
            else:
                # Fallback al método original si Barranquilla falla
                results["temp"] = [float(x) if x is not None else None for x in temp_complete]
                results["max_temp"] = [float(x) if x is not None else None for x in temp_max_complete]
                results["min_temp"] = [float(x) if x is not None else None for x in temp_min_complete]
            
            results["total_predictions"] = float(sum(current_month_demands) + sum(predictions))
            
            results["clima"] = clima_ciudades

            return results
        except Exception as e:
            raise Exception(f"Exception on predict: {str(e)}")
    
    def generate_comparison_excel(self, search: MPMPredictSchema, db: Session):
        """
        Genera un Excel comparando los 4 métodos de predicción MPM
        
        Args:
            search: Esquema con los parámetros de búsqueda
            db: Sesión de base de datos
            
        Returns:
            io.BytesIO: Buffer con el archivo Excel
        """
        try:
            return MPMExcelGenerator.generate_comparison_excel(self, search, db)
        except Exception as e:
            raise Exception(f"Error al generar Excel de comparación: {str(e)}")

class MPMExcelGenerator:
    @staticmethod
    def generate_comparison_excel(mpm_service, search_base: MPMPredictSchema, db: Session):
        """
        Genera un archivo Excel con las predicciones de los 4 métodos MPM
        
        Args:
            mmp_service: Instancia de MPMService
            search_base: Esquema base con los parámetros comunes de búsqueda
            db: Sesión de base de datos
            
        Returns:
            io.BytesIO: Buffer con el archivo Excel
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import io
            from datetime import datetime
            
            # ===== DICCIONARIOS DE TRADUCCIÓN AL ESPAÑOL =====
            MESES_ESPANOL = {
                1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
                5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto", 
                9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
            }
            
            DIAS_ESPANOL = {
                0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves",
                4: "Viernes", 5: "Sábado", 6: "Domingo"
            }
            
            print(f"Generando Excel con parámetros: last_date={search_base.last_date}, previous_days={search_base.previous_days}")
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Predicciones_MPM"
            
            # Configurar estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Obtener predicciones de los 4 métodos
            predictions_data = {}
            method_names = {
                1: "Solo Demanda",
                2: "Una Temperatura", 
                3: "Tres Temperaturas",
                4: "Promedio"
            }
            
            # Para cada método, obtener las predicciones
            for method_id in [1, 2, 3, 4]:
                search = MPMPredictSchema(
                    last_date=search_base.last_date,
                    previous_days=search_base.previous_days,
                    id=method_id,
                    type=search_base.type if method_id in [2, 3] else 0  # Usar 0 como valor por defecto
                )
                
                try:
                    # CORRECCIÓN: Cambiar mmp_service por mpm_service
                    result = mpm_service.predict(search, db)
                    predictions_data[method_id] = result
                    print(f"Método {method_id} - Datos obtenidos correctamente")
                    if result:
                        print(f"  - Fechas: {len(result.get('dates', []))}")
                        print(f"  - Históricos: {len(result.get('historic_demands', []))}")
                        print(f"  - Predicciones: {len(result.get('prediction', []))}")
                except Exception as e:
                    print(f"Error al obtener predicciones del método {method_id}: {str(e)}")
                    predictions_data[method_id] = None
            
            # Verificar que se obtuvieron datos de al menos un método
            valid_methods = [method_id for method_id, data in predictions_data.items() if data is not None]
            if not valid_methods:
                print("ERROR: No se pudieron obtener datos de ningún método de predicción")
                # Crear Excel con mensaje de error en español
                ws['A10'] = "ERROR: No se pudieron obtener datos de predicción de ningún método"
                ws['A11'] = "Verifique los parámetros de entrada y la conectividad a la base de datos"
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                return excel_buffer
            
            print(f"Métodos válidos obtenidos: {valid_methods}")
            
            # Obtener las fechas y datos
            all_dates = []
            
            # Buscar el primer método que tenga datos válidos para obtener las fechas
            for method_id, data in predictions_data.items():
                if data and 'dates' in data and len(data['dates']) > 0:
                    all_dates = data['dates']
                    print(f"Fechas obtenidas del método {method_id}: {len(all_dates)} fechas")
                    break
            
            if not all_dates:
                print("No se encontraron fechas en ningún método")
                # Crear Excel vacío con mensaje de error en español
                ws['A10'] = "ERROR: No se pudieron obtener datos de predicción"
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                return excel_buffer
            
            # Crear estructura del Excel
            # Título principal
            ws['A1'] = "COMPARACIÓN DE MÉTODOS DE PREDICCIÓN MPM"
            ws.merge_cells('A1:G1')
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = center_alignment
            
            # ===== INFORMACIÓN DEL PERÍODO EN ESPAÑOL =====
            last_date = datetime.strptime(search_base.last_date, "%Y-%m-%d")
            
            # Formatear fecha en español
            mes_numero = last_date.month
            ano = last_date.year
            mes_español = MESES_ESPANOL[mes_numero]
            
            ws['A3'] = f"Mes: {mes_español} {ano}"
            ws['A4'] = f"Última fecha con datos: {search_base.last_date}"
            ws['A5'] = f"Días previos considerados: {search_base.previous_days}"
            
            # Headers de la tabla principal en español
            row_start = 7
            headers = [
                "Fecha", 
                "Histórico/Predicción", 
                "Método 1\nSolo Demanda\n(MW)", 
                "Método 2\nUna Temperatura\n(MW)", 
                "Método 3\nTres Temperaturas\n(MW)", 
                "Método 4\nPromedio\n(MW)", 
                "Promedio\nMétodos\n(MW)"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_start, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            
            # Llenar los datos
            current_row = row_start + 1
            daily_averages = []
            
            # Factor de conversión: 1 MW = 1,000,000 W
            WATTS_TO_MW = 1_000_000
            
            # Obtener la cantidad de datos históricos del primer método válido
            historic_count = 0
            for method_id, data in predictions_data.items():
                if data and 'historic_demands' in data:
                    historic_count = len(data['historic_demands'])
                    break
            
            print(f"Total fechas: {len(all_dates)}, Históricos: {historic_count}")
            
            for i, date in enumerate(all_dates):
                # Fecha
                ws.cell(row=current_row, column=1, value=date)
                
                # Determinar si es histórico o predicción
                is_historic = i < historic_count
                
                # Tipo de dato en español
                ws.cell(row=current_row, column=2, value="Histórico" if is_historic else "Predicción")
                
                # Valores de cada método convertidos a MW
                method_values = []
                for method_id in [1, 2, 3, 4]:
                    col = method_id + 2  # +2 porque ahora tenemos una columna extra
                    value_in_watts = None
                    
                    if predictions_data[method_id]:
                        try:
                            if is_historic:
                                # Obtener dato histórico
                                historic_data = predictions_data[method_id].get('historic_demands', [])
                                if i < len(historic_data):
                                    value_in_watts = historic_data[i]
                            else:
                                # Obtener predicción
                                prediction_data = predictions_data[method_id].get('prediction', [])
                                prediction_index = i - historic_count
                                if 0 <= prediction_index < len(prediction_data):
                                    value_in_watts = prediction_data[prediction_index]
                        except (IndexError, KeyError) as e:
                            print(f"Error accediendo a datos del método {method_id}, índice {i}: {e}")
                            value_in_watts = None
                    
                    if value_in_watts is not None:
                        try:
                            value_in_mw = float(value_in_watts) / WATTS_TO_MW
                            ws.cell(row=current_row, column=col, value=value_in_mw)
                            method_values.append(value_in_mw)
                        except (ValueError, TypeError) as e:
                            print(f"Error convirtiendo valor {value_in_watts} a MW: {e}")
                            ws.cell(row=current_row, column=col, value="ERROR")
                    else:
                        ws.cell(row=current_row, column=col, value="N/D")  # "No Disponible" en español
                
                # Promedio de los métodos en MW
                if method_values:
                    avg_value = sum(method_values) / len(method_values)
                    ws.cell(row=current_row, column=7, value=avg_value)
                    daily_averages.append(avg_value)
                else:
                    ws.cell(row=current_row, column=7, value="N/D")  # "No Disponible" en español
                
                current_row += 1
            
            # Fila de totales en español
            total_row = current_row + 1
            ws.cell(row=total_row, column=1, value="TOTAL (MW)")
            ws.cell(row=total_row, column=1).font = Font(bold=True)
            
            # Calcular totales para cada método en MW
            for method_id in [1, 2, 3, 4]:
                col = method_id + 2  # +2 por la columna extra
                try:
                    if predictions_data[method_id] and 'total_predictions' in predictions_data[method_id]:
                        total_in_watts = float(predictions_data[method_id]['total_predictions'])
                        total_in_mw = total_in_watts / WATTS_TO_MW
                        ws.cell(row=total_row, column=col, value=total_in_mw)
                        print(f"Total método {method_id}: {total_in_mw} MW")
                    else:
                        ws.cell(row=total_row, column=col, value="N/D")  # "No Disponible" en español
                        print(f"Método {method_id}: Sin datos de total")
                except (ValueError, TypeError, KeyError) as e:
                    print(f"Error calculando total del método {method_id}: {e}")
                    ws.cell(row=total_row, column=col, value="ERROR")
            
            # Total del promedio
            try:
                if daily_averages and len(daily_averages) > 0:
                    total_average = sum(daily_averages)
                    ws.cell(row=total_row, column=7, value=total_average)
                    print(f"Total promedio: {total_average} MW")
                else:
                    ws.cell(row=total_row, column=7, value="N/D")  # "No Disponible" en español
                    print("No hay datos para calcular el promedio total")
            except Exception as e:
                print(f"Error calculando total del promedio: {e}")
                ws.cell(row=total_row, column=7, value="ERROR")
            
            # Aplicar formato a todas las celdas con datos
            for row in range(row_start + 1, total_row + 1):
                for col in range(1, 8):  # Ahora son 7 columnas
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = center_alignment
                    
                    # Formato numérico para las columnas de valores
                    if col > 2 and isinstance(cell.value, (int, float)):  # Col > 2 por las nuevas columnas
                        cell.number_format = '#,##0.00'
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            for col in range(3, 8):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # ===== AGREGAR INFORMACIÓN ADICIONAL EN ESPAÑOL =====
            # Agregar nota explicativa
            note_row = total_row + 3
            ws.cell(row=note_row, column=1, value="NOTAS:")
            ws.cell(row=note_row, column=1).font = Font(bold=True)
            
            ws.cell(row=note_row + 1, column=1, value="• Método 1: Predicción basada únicamente en datos históricos de demanda")
            ws.cell(row=note_row + 2, column=1, value="• Método 2: Predicción con datos de demanda y una temperatura promedio")
            ws.cell(row=note_row + 3, column=1, value="• Método 3: Predicción con demanda y temperaturas de tres ciudades (Bolívar, Barranquilla, Cesar)")
            ws.cell(row=note_row + 4, column=1, value="• Método 4: Predicción basada en promedios por tipo de día (Laboral/Sábado/Domingo-Festivo)")
            ws.cell(row=note_row + 5, column=1, value="• Los valores están expresados en Megavatios (MW)")
            
            # Ajustar ancho para las notas
            ws.column_dimensions['A'].width = 80
            
            # Agregar fecha de generación del reporte
            fecha_generacion = datetime.now()
            mes_gen = MESES_ESPANOL[fecha_generacion.month]
            ws.cell(row=note_row + 7, column=1, value=f"Reporte generado el: {fecha_generacion.day} de {mes_gen} de {fecha_generacion.year} a las {fecha_generacion.strftime('%H:%M')}")
            ws.cell(row=note_row + 7, column=1).font = Font(italic=True, size=10)
            
            # Guardar en buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            print("Excel generado exitosamente en español")
            return excel_buffer

        except Exception as e:
            print(f"Error general en generate_comparison_excel: {str(e)}")
            import traceback
            print(f"Traceback: {traceback.format_exc()}")
            
            # Crear Excel básico con mensaje de error en español
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Error"
                ws['A1'] = "ERROR AL GENERAR REPORTE"
                ws['A2'] = f"Error: {str(e)}"
                ws['A3'] = f"Fecha: {search_base.last_date}"
                ws['A4'] = f"Días previos: {search_base.previous_days}"
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                return excel_buffer
            except:
                # Si incluso esto falla, devolver buffer vacío
                return io.BytesIO()
